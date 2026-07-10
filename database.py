import os
import sys
import json
import sqlite3
from contextlib import contextmanager
from datetime import datetime, date
from utils import (get_meeting_datetimes, get_data_dir, migrar_dados_antigos,
                   DEPARTAMENTOS_RECEBIDOS, iso_to_display)

# Dados persistentes ficam em %LOCALAPPDATA%\GestaoDocumentosDNE (fora da
# pasta de instalação), para sobreviver a reconstruções do executável.
_BASE_DIR = get_data_dir()
if getattr(sys, 'frozen', False):
    migrar_dados_antigos(_BASE_DIR, os.path.dirname(sys.executable))

DB_PATH = os.path.join(_BASE_DIR, 'gestao_documentos.db')


class Database:
    def __init__(self, db_path=None):
        # db_path opcional permite usar uma base de dados alternativa (ex: testes)
        self.db_path = db_path or DB_PATH
        self.init_db()

    @contextmanager
    def _connect(self):
        """Context manager que abre a ligação, faz commit em caso de sucesso e
        garante o fecho da ligação mesmo que ocorra uma excepção. Evita ligações
        penduradas (que no Windows podem bloquear o ficheiro da base de dados)."""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
            conn.commit()
        finally:
            conn.close()

    def get_connection(self):
        """Mantido por compatibilidade. Prefira o context manager _connect()."""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def init_db(self):
        with self._connect() as conn:
            c = conn.cursor()

            c.execute('''CREATE TABLE IF NOT EXISTS documentos_recebidos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero TEXT NOT NULL,
                proveniencia TEXT,
                remetente_nome TEXT,
                remetente_cargo TEXT,
                assunto TEXT NOT NULL,
                data_recepcao TEXT,
                despacho TEXT,
                endereçado_a TEXT,
                tecnico TEXT,
                data_resposta TEXT,
                prazo_status TEXT DEFAULT 'Pendente',
                prazo_dias INTEGER,
                prazo_data TEXT,
                observacao TEXT,
                ficheiro_path TEXT,
                ficheiro_resposta_path TEXT,
                created_at TEXT DEFAULT (datetime('now'))
            )''')

            c.execute('''CREATE TABLE IF NOT EXISTS documentos_enviados (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero TEXT,
                assunto TEXT NOT NULL,
                preparado_por TEXT,
                assinante TEXT,
                destinatario_nome TEXT,
                destinatario_cargo TEXT,
                instituicao TEXT,
                data_envio TEXT,
                ficheiro_path TEXT,
                observacao TEXT,
                created_at TEXT DEFAULT (datetime('now'))
            )''')

            c.execute('''CREATE TABLE IF NOT EXISTS reunioes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                num_doc TEXT,
                organizador TEXT,
                data_convocatoria TEXT,
                assunto TEXT NOT NULL,
                data_reuniao TEXT,
                hora_local TEXT,
                link_convocatoria TEXT,
                participantes TEXT,
                contactos TEXT,
                decisoes TEXT,
                ficheiro_path TEXT,
                created_at TEXT DEFAULT (datetime('now'))
            )''')

            c.execute('''CREATE TABLE IF NOT EXISTS contactos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero INTEGER,
                nome TEXT NOT NULL,
                email TEXT,
                telefone TEXT,
                departamento TEXT,
                cargo TEXT,
                created_at TEXT DEFAULT (datetime('now'))
            )''')

            # Reciclagem: registos eliminados ficam aqui 30 dias e podem ser
            # restaurados nas Configurações antes da purga automática.
            c.execute('''CREATE TABLE IF NOT EXISTS reciclagem (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tabela TEXT NOT NULL,
                registo_id INTEGER,
                descricao TEXT,
                dados TEXT NOT NULL,
                eliminado_em TEXT DEFAULT (datetime('now', 'localtime'))
            )''')
            # Purga automática: itens eliminados há mais de 30 dias
            c.execute("DELETE FROM reciclagem "
                      "WHERE eliminado_em < datetime('now', 'localtime', '-30 days')")

            # ── Migração: garante novas colunas em bases de dados existentes ───────
            c.execute("PRAGMA table_info(documentos_recebidos)")
            cols_recebidos = {row[1] for row in c.fetchall()}
            if 'ficheiro_resposta_path' not in cols_recebidos:
                c.execute("ALTER TABLE documentos_recebidos ADD COLUMN ficheiro_resposta_path TEXT")
            if 'prazo_dias' not in cols_recebidos:
                c.execute("ALTER TABLE documentos_recebidos ADD COLUMN prazo_dias INTEGER")
            if 'prazo_data' not in cols_recebidos:
                c.execute("ALTER TABLE documentos_recebidos ADD COLUMN prazo_data TEXT")

            c.execute("PRAGMA table_info(reunioes)")
            cols_reunioes = {row[1] for row in c.fetchall()}
            if 'cancelada' not in cols_reunioes:
                c.execute("ALTER TABLE reunioes ADD COLUMN cancelada INTEGER DEFAULT 0")

            # Índices para acelerar filtros, relatórios e ordenações
            c.execute("CREATE INDEX IF NOT EXISTS idx_rec_data ON documentos_recebidos(data_recepcao)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_rec_status ON documentos_recebidos(prazo_status)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_rec_tecnico ON documentos_recebidos(tecnico)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_env_data ON documentos_enviados(data_envio)")
            c.execute("CREATE INDEX IF NOT EXISTS idx_reu_data ON reunioes(data_reuniao)")

            # Corrige registos antigos com o nome de departamento mal escrito
            c.execute(
                "UPDATE documentos_recebidos SET endereçado_a = 'Dep. de Planeamento Energético' "
                "WHERE endereçado_a = 'Dep. de Planeamento Enegético'"
            )

            # Load initial data only if tables are empty
            c.execute("SELECT COUNT(*) FROM contactos")
            if c.fetchone()[0] == 0:
                self._load_initial_contactos(c)

    def _load_initial_contactos(self, c):
        records = [
            (1, 'Dir. Marcelina Mataveia', 'mmataveia@yahoo.com', '824195400 / 840495452', 'Direcção', 'Directora'),
            (2, 'Dir. Ortigio Nhanombe', 'ortigiolois@gmail.com', '823959740', 'Direcção', 'Director'),
            (3, 'Iazalde Jose Jeremias', 'iazaldejose@gmail.com', '842821959', 'Dep. Planeamento Energético', 'Técnico'),
            (4, 'Estacio Chumbitico', 'estaciochumbitico@gmail.com', '840409362', 'Dep. Planeamento Energético', 'Técnico'),
            (5, 'Luis Guambe', 'luisguambe84@gmail.com', '848224897', 'Dep. Planeamento Energético', 'Técnico'),
            (6, 'Armindo Jone', 'armindojone@gmail.com', '846691153', 'Dep. Planeamento Energético', 'Técnico'),
            (7, 'Sergio Honwana', 'gabarsingue@gmail.com', '824099656', 'Dep. Planeamento Energético', 'Técnico'),
            (8, 'Anisio Pinto', 'anisiopintomanuel@gmail.com', '848576118', 'Dep. Planeamento Energético', 'Técnico'),
            (9, 'Bernardo Joao Nopia', 'nopiha20@yahoo.com.br', '848285831', 'Dep. Estudos e Projectos', 'Técnico'),
            (10, 'Francisco Mahangue', 'mahangue.fr@gmail.com', '842555253', 'Dep. Estudos e Projectos', 'Técnico'),
            (11, 'Manuel Andissene', 'manuelandissene@gmail.com', '829674910', 'Dep. Estudos e Projectos', 'Técnico'),
            (12, 'Rogerio Manhica', 'rogerio.manhica@gmail.com', '846733413', 'Dep. Licenciamento e Fiscalização', 'Técnico'),
            (13, 'Jose Mapilele', 'mapilas.ze@gmail.com', '843535174', 'Dep. Eficiência Energética', 'Técnico'),
            (14, 'Damiao Namuera', 'dnamuera@gmail.com', '824166198', 'Dep. Eficiência Energética', 'Técnico'),
            (15, 'Issufo Juma', 'issufojuma2003@yahoo.com.br', '845082416', 'Dep. Energias Renováveis', 'Técnico'),
            (16, 'Cristiano Gumete', 'cristianogumete@gmail.com', '843773052', 'Dep. Energias Renováveis', 'Técnico'),
            (17, 'Victoria Safrao', 'victoriasafrao@gmail.com', '845351745', 'Dep. Energias Renováveis', 'Técnico'),
            (18, 'Teresa Estevao', 'dne.mireme@mireme.gov.mz', '865611270', 'Rep. Administração e Finanças', 'Técnico'),
            (19, 'Benvinda Tembe', 'benvindateembe@gmail.com', '846582143', 'Rep. Administração e Finanças', 'Técnico'),
            (20, 'Susana Gomez', 'S.Gomez@institute.global', '845210770', 'Transição Energética', 'Técnico'),
            (21, 'Inocencio Gujamo', 'inocencio.gujamo@gmail.com', '844777042', 'UIPCE', 'Técnico'),
        ]
        c.executemany(
            'INSERT INTO contactos (numero, nome, email, telefone, departamento, cargo) VALUES (?,?,?,?,?,?)',
            records
        )

    # ---- Autocomplete ----
    def get_autocomplete(self, campo):
        """Devolve lista ordenada de valores únicos previamente introduzidos num campo."""
        _MAP = {
            'proveniencia':    ('documentos_recebidos', 'proveniencia'),
            'remetente_nome':  ('documentos_recebidos', 'remetente_nome'),
            'remetente_cargo': ('documentos_recebidos', 'remetente_cargo'),
            'tecnico':         ('documentos_recebidos', 'tecnico'),
            'destinatario_nome':  ('documentos_enviados', 'destinatario_nome'),
            'destinatario_cargo': ('documentos_enviados', 'destinatario_cargo'),
            'instituicao':        ('documentos_enviados', 'instituicao'),
            'assinante':          ('documentos_enviados', 'assinante'),
            'preparado_por':      ('documentos_enviados', 'preparado_por'),
        }
        if campo not in _MAP:
            return []
        tabela, coluna = _MAP[campo]
        with self._connect() as conn:
            c = conn.cursor()
            c.execute(f"SELECT DISTINCT {coluna} FROM {tabela} "
                      f"WHERE {coluna} IS NOT NULL AND {coluna} != '' ORDER BY {coluna}")
            return [r[0] for r in c.fetchall()]

    def lookup_remetente(self, nome):
        """Dado o nome de um remetente, devolve {'remetente_cargo', 'proveniencia'}
        com os valores mais recentes não-vazios já registados em Recebidos para
        esse nome (comparação sem distinção de maiúsculas/acentos de espaços).
        Devolve {} se o nome não constar. Usado para preenchimento automático."""
        alvo = (nome or '').strip().casefold()
        if not alvo:
            return {}
        with self._connect() as conn:
            c = conn.cursor()
            c.execute("SELECT remetente_nome, remetente_cargo, proveniencia "
                      "FROM documentos_recebidos ORDER BY id DESC")
            rows = c.fetchall()
        res = {}
        for rn, cargo, prov in rows:
            if (rn or '').strip().casefold() != alvo:
                continue
            if 'remetente_cargo' not in res and cargo and cargo.strip():
                res['remetente_cargo'] = cargo.strip()
            if 'proveniencia' not in res and prov and prov.strip():
                res['proveniencia'] = prov.strip()
            if len(res) == 2:
                break
        return res

    def lookup_destinatario(self, nome):
        """Dado o nome de um destinatário, devolve {'destinatario_cargo',
        'instituicao'} com os valores mais recentes não-vazios já registados em
        Enviados para esse nome. Devolve {} se não constar. Preenchimento automático."""
        alvo = (nome or '').strip().casefold()
        if not alvo:
            return {}
        with self._connect() as conn:
            c = conn.cursor()
            c.execute("SELECT destinatario_nome, destinatario_cargo, instituicao "
                      "FROM documentos_enviados ORDER BY id DESC")
            rows = c.fetchall()
        res = {}
        for dn, cargo, inst in rows:
            if (dn or '').strip().casefold() != alvo:
                continue
            if 'destinatario_cargo' not in res and cargo and cargo.strip():
                res['destinatario_cargo'] = cargo.strip()
            if 'instituicao' not in res and inst and inst.strip():
                res['instituicao'] = inst.strip()
            if len(res) == 2:
                break
        return res

    def find_numero_duplicado(self, tabela, numero, excluir_id=None):
        """Devolve o assunto de um documento já registado com este número
        (excluindo o próprio registo em edição), ou None se não existir.
        Evita carregar a tabela inteira só para validar duplicados."""
        table = 'documentos_recebidos' if tabela == 'recebidos' else 'documentos_enviados'
        with self._connect() as conn:
            c = conn.cursor()
            c.execute(f"SELECT assunto FROM {table} WHERE numero=? AND id<>? LIMIT 1",
                      (numero, excluir_id if excluir_id is not None else -1))
            row = c.fetchone()
            return row[0] if row else None

    # ---- Recebidos ----
    def get_all_recebidos(self, filters=None):
        with self._connect() as conn:
            c = conn.cursor()
            query = "SELECT * FROM documentos_recebidos WHERE 1=1"
            params = []
            if filters:
                if filters.get('search'):
                    s = f"%{filters['search']}%"
                    query += " AND (numero LIKE ? OR assunto LIKE ? OR remetente_nome LIKE ? OR proveniencia LIKE ? OR observacao LIKE ?)"
                    params += [s, s, s, s, s]
                if filters.get('tecnico'):
                    query += " AND tecnico = ?"
                    params.append(filters['tecnico'])
                if filters.get('prazo_status'):
                    query += " AND prazo_status = ?"
                    params.append(filters['prazo_status'])
                if filters.get('data_inicio'):
                    query += " AND data_recepcao >= ?"
                    params.append(filters['data_inicio'])
                if filters.get('data_fim'):
                    query += " AND data_recepcao <= ?"
                    params.append(filters['data_fim'])
            query += " ORDER BY data_recepcao DESC"
            c.execute(query, params)
            return [dict(r) for r in c.fetchall()]

    def get_recebido(self, id):
        with self._connect() as conn:
            c = conn.cursor()
            c.execute("SELECT * FROM documentos_recebidos WHERE id=?", (id,))
            row = c.fetchone()
            return dict(row) if row else None

    def insert_recebido(self, data):
        with self._connect() as conn:
            c = conn.cursor()
            c.execute('''INSERT INTO documentos_recebidos
                (numero, proveniencia, remetente_nome, remetente_cargo, assunto, data_recepcao,
                 despacho, endereçado_a, tecnico, data_resposta, prazo_status, prazo_data, observacao, ficheiro_path,
                 ficheiro_resposta_path)
                VALUES (:numero, :proveniencia, :remetente_nome, :remetente_cargo, :assunto, :data_recepcao,
                        :despacho, :endereçado_a, :tecnico, :data_resposta, :prazo_status, :prazo_data, :observacao, :ficheiro_path,
                        :ficheiro_resposta_path)''',
                      data)
            return c.lastrowid

    def update_recebido(self, id, data):
        with self._connect() as conn:
            c = conn.cursor()
            c.execute('''UPDATE documentos_recebidos SET
                numero=:numero, proveniencia=:proveniencia, remetente_nome=:remetente_nome,
                remetente_cargo=:remetente_cargo, assunto=:assunto, data_recepcao=:data_recepcao,
                despacho=:despacho, endereçado_a=:endereçado_a, tecnico=:tecnico,
                data_resposta=:data_resposta, prazo_status=:prazo_status, prazo_data=:prazo_data,
                observacao=:observacao, ficheiro_path=:ficheiro_path,
                ficheiro_resposta_path=:ficheiro_resposta_path
                WHERE id=:id''',
                      {**data, 'id': id})
        return True

    def delete_recebido(self, id):
        return self._mover_para_reciclagem('documentos_recebidos', id,
                                           ('numero', 'assunto'))

    def recalcular_prazos(self, prazo_padrao=5, usar_uteis=False):
        """Recalcula o Status Prazo de todos os documentos recebidos comparando a
        Data de Resposta com a data-limite de cada documento. A data-limite é a
        específica do documento (prazo_data), ou, se vazia, a Data de Recepção
        mais o prazo padrão global. Documentos marcados manualmente como
        'Arquivado' ou 'Arquivo' não são alterados."""
        from utils import data_limite
        with self._connect() as conn:
            c = conn.cursor()
            c.execute("SELECT id, data_recepcao, data_resposta, prazo_status, prazo_data FROM documentos_recebidos")
            rows = c.fetchall()
            for r in rows:
                status_actual = r['prazo_status']
                if status_actual in ('Arquivado', 'Arquivo'):
                    continue
                if not r['data_resposta']:
                    novo = 'Pendente'
                else:
                    limite = data_limite(r['data_recepcao'], r['prazo_data'], prazo_padrao, usar_uteis)
                    if not limite:
                        continue
                    try:
                        d2 = datetime.strptime(r['data_resposta'], '%Y-%m-%d').date()
                        d1 = datetime.strptime(r['data_recepcao'], '%Y-%m-%d').date()
                        dl = datetime.strptime(limite, '%Y-%m-%d').date()
                        if d2 < d1:
                            continue
                        novo = 'Dentro do Prazo' if d2 <= dl else 'Fora do Prazo'
                    except Exception:
                        continue
                if novo != status_actual:
                    c.execute("UPDATE documentos_recebidos SET prazo_status=? WHERE id=?", (novo, r['id']))

    # ---- Enviados ----
    def get_all_enviados(self, filters=None):
        with self._connect() as conn:
            c = conn.cursor()
            query = "SELECT * FROM documentos_enviados WHERE 1=1"
            params = []
            if filters:
                if filters.get('search'):
                    s = f"%{filters['search']}%"
                    query += " AND (numero LIKE ? OR assunto LIKE ? OR destinatario_nome LIKE ? OR instituicao LIKE ? OR observacao LIKE ?)"
                    params += [s, s, s, s, s]
                if filters.get('assinante'):
                    query += " AND assinante = ?"
                    params.append(filters['assinante'])
                if filters.get('preparado_por'):
                    query += " AND preparado_por = ?"
                    params.append(filters['preparado_por'])
                if filters.get('data_inicio'):
                    query += " AND data_envio >= ?"
                    params.append(filters['data_inicio'])
                if filters.get('data_fim'):
                    query += " AND data_envio <= ?"
                    params.append(filters['data_fim'])
            query += " ORDER BY data_envio DESC"
            c.execute(query, params)
            return [dict(r) for r in c.fetchall()]

    def get_enviado(self, id):
        with self._connect() as conn:
            c = conn.cursor()
            c.execute("SELECT * FROM documentos_enviados WHERE id=?", (id,))
            row = c.fetchone()
            return dict(row) if row else None

    def insert_enviado(self, data):
        with self._connect() as conn:
            c = conn.cursor()
            c.execute('''INSERT INTO documentos_enviados
                (numero, assunto, preparado_por, assinante, destinatario_nome, destinatario_cargo,
                 instituicao, data_envio, ficheiro_path, observacao)
                VALUES (:numero, :assunto, :preparado_por, :assinante, :destinatario_nome,
                        :destinatario_cargo, :instituicao, :data_envio, :ficheiro_path, :observacao)''',
                      data)
            return c.lastrowid

    def update_enviado(self, id, data):
        with self._connect() as conn:
            c = conn.cursor()
            c.execute('''UPDATE documentos_enviados SET
                numero=:numero, assunto=:assunto, preparado_por=:preparado_por, assinante=:assinante,
                destinatario_nome=:destinatario_nome, destinatario_cargo=:destinatario_cargo,
                instituicao=:instituicao, data_envio=:data_envio, ficheiro_path=:ficheiro_path,
                observacao=:observacao WHERE id=:id''',
                      {**data, 'id': id})
        return True

    def delete_enviado(self, id):
        return self._mover_para_reciclagem('documentos_enviados', id,
                                           ('numero', 'assunto'))

    # ---- Reunioes ----
    def get_all_reunioes(self, filters=None):
        with self._connect() as conn:
            c = conn.cursor()
            query = "SELECT * FROM reunioes WHERE 1=1"
            params = []
            if filters:
                if filters.get('search'):
                    s = f"%{filters['search']}%"
                    query += (" AND (assunto LIKE ? OR organizador LIKE ? OR num_doc LIKE ?"
                              " OR participantes LIKE ? OR hora_local LIKE ?)")
                    params += [s, s, s, s, s]
                if filters.get('data_reuniao'):
                    query += " AND data_reuniao = ?"
                    params.append(filters['data_reuniao'])
            query += " ORDER BY data_reuniao ASC"
            c.execute(query, params)
            return [dict(r) for r in c.fetchall()]

    def get_reuniao(self, id):
        with self._connect() as conn:
            c = conn.cursor()
            c.execute("SELECT * FROM reunioes WHERE id=?", (id,))
            row = c.fetchone()
            return dict(row) if row else None

    def insert_reuniao(self, data):
        with self._connect() as conn:
            c = conn.cursor()
            c.execute('''INSERT INTO reunioes
                (num_doc, organizador, data_convocatoria, assunto, data_reuniao, hora_local,
                 link_convocatoria, participantes, contactos, decisoes, ficheiro_path, cancelada)
                VALUES (:num_doc, :organizador, :data_convocatoria, :assunto, :data_reuniao,
                        :hora_local, :link_convocatoria, :participantes, :contactos, :decisoes,
                        :ficheiro_path, :cancelada)''',
                      data)
            return c.lastrowid

    def update_reuniao(self, id, data):
        with self._connect() as conn:
            c = conn.cursor()
            c.execute('''UPDATE reunioes SET
                num_doc=:num_doc, organizador=:organizador, data_convocatoria=:data_convocatoria,
                assunto=:assunto, data_reuniao=:data_reuniao, hora_local=:hora_local,
                link_convocatoria=:link_convocatoria, participantes=:participantes,
                contactos=:contactos, decisoes=:decisoes, ficheiro_path=:ficheiro_path,
                cancelada=:cancelada
                WHERE id=:id''',
                      {**data, 'id': id})
        return True

    def delete_reuniao(self, id):
        return self._mover_para_reciclagem('reunioes', id,
                                           ('num_doc', 'assunto'))

    # ---- Contactos ----
    def get_all_contactos(self, search=None):
        with self._connect() as conn:
            c = conn.cursor()
            if search:
                s = f"%{search}%"
                c.execute("SELECT * FROM contactos WHERE nome LIKE ? OR departamento LIKE ? "
                          "OR email LIKE ? OR telefone LIKE ? OR cargo LIKE ? ORDER BY numero",
                          (s, s, s, s, s))
            else:
                c.execute("SELECT * FROM contactos ORDER BY numero")
            return [dict(r) for r in c.fetchall()]

    def get_contacto(self, id):
        with self._connect() as conn:
            c = conn.cursor()
            c.execute("SELECT * FROM contactos WHERE id=?", (id,))
            row = c.fetchone()
            return dict(row) if row else None

    def insert_contacto(self, data):
        with self._connect() as conn:
            c = conn.cursor()
            c.execute('''INSERT INTO contactos (numero, nome, email, telefone, departamento, cargo)
                         VALUES (:numero, :nome, :email, :telefone, :departamento, :cargo)''', data)
            return c.lastrowid

    def update_contacto(self, id, data):
        with self._connect() as conn:
            c = conn.cursor()
            c.execute('''UPDATE contactos SET numero=:numero, nome=:nome, email=:email,
                         telefone=:telefone, departamento=:departamento, cargo=:cargo WHERE id=:id''',
                      {**data, 'id': id})
        return True

    def delete_contacto(self, id):
        return self._mover_para_reciclagem('contactos', id, ('nome', 'cargo'))

    def get_nomes_contactos(self):
        with self._connect() as conn:
            c = conn.cursor()
            c.execute("SELECT nome FROM contactos ORDER BY nome")
            return [r[0] for r in c.fetchall()]

    # ---- Relatorio ----
    def get_anos_disponiveis(self):
        """Devolve lista de anos com documentos registados, ordenados DESC."""
        with self._connect() as conn:
            c = conn.cursor()
            c.execute("""SELECT DISTINCT substr(data_recepcao, 1, 4) as ano
                         FROM documentos_recebidos WHERE data_recepcao IS NOT NULL AND data_recepcao != ''
                         UNION
                         SELECT DISTINCT substr(data_envio, 1, 4) as ano
                         FROM documentos_enviados WHERE data_envio IS NOT NULL AND data_envio != ''
                         ORDER BY ano DESC""")
            anos = [r[0] for r in c.fetchall() if r[0]]
        ano_atual = str(date.today().year)
        if ano_atual not in anos:
            anos.insert(0, ano_atual)
        return anos

    def get_relatorio_stats(self, ano=None, mes=None):
        with self._connect() as conn:
            c = conn.cursor()
            ano = ano or str(date.today().year)

            # prefixo de data para filtro: "2026-03" (ano+mês) ou "2026" (só ano)
            if mes and mes != "0":
                prefixo_rec = f"{ano}-{int(mes):02d}"
                prefixo_env = prefixo_rec
                prefixo_reu = prefixo_rec
            else:
                prefixo_rec = ano
                prefixo_env = ano
                prefixo_reu = date.today().strftime('%Y-%m') if ano == str(date.today().year) else f"{ano}-01"

            c.execute("SELECT COUNT(*) FROM documentos_recebidos WHERE data_recepcao LIKE ?", (f"{prefixo_rec}%",))
            total_recebidos = c.fetchone()[0]

            c.execute("""SELECT COUNT(*) FROM documentos_recebidos
                         WHERE data_resposta IS NOT NULL AND data_resposta != ''
                         AND data_recepcao LIKE ?""", (f"{prefixo_rec}%",))
            total_respondidos = c.fetchone()[0]

            c.execute("""SELECT COUNT(*) FROM documentos_recebidos
                         WHERE prazo_status IN ('Dentro do Prazo', 'Arquivado', 'Arquivo')
                         AND data_recepcao LIKE ?""", (f"{prefixo_rec}%",))
            total_dentro = c.fetchone()[0]

            c.execute("""SELECT COUNT(*) FROM documentos_recebidos
                         WHERE prazo_status='Fora do Prazo' AND data_recepcao LIKE ?""", (f"{prefixo_rec}%",))
            total_fora = c.fetchone()[0]

            c.execute("SELECT COUNT(*) FROM reunioes WHERE data_reuniao LIKE ?", (f"{prefixo_reu}%",))
            reunioes_mes = c.fetchone()[0]

            c.execute("SELECT COUNT(*) FROM documentos_enviados WHERE data_envio LIKE ?", (f"{prefixo_env}%",))
            total_enviados = c.fetchone()[0]

        respondidos_com_status = total_dentro + total_fora
        taxa = round((total_dentro / respondidos_com_status * 100) if respondidos_com_status > 0 else 0, 1)

        return {
            'total_recebidos': total_recebidos,
            'total_respondidos': total_respondidos,
            'taxa_cumprimento': taxa,
            'total_fora_prazo': total_fora,
            'reunioes_mes': reunioes_mes,
            'total_enviados': total_enviados,
        }

    def get_relatorio_departamentos(self, ano=None, mes=None):
        with self._connect() as conn:
            c = conn.cursor()
            ano = ano or str(date.today().year)
            if mes and mes != "0":
                prefixo = f"{ano}-{int(mes):02d}"
            else:
                prefixo = ano

            # Uma única consulta agregada (em vez de 3 por departamento)
            c.execute("""SELECT endereçado_a AS dep,
                                COUNT(*) AS total,
                                SUM(CASE WHEN prazo_status IN ('Dentro do Prazo', 'Arquivado', 'Arquivo')
                                         THEN 1 ELSE 0 END) AS dentro,
                                SUM(CASE WHEN prazo_status='Fora do Prazo' THEN 1 ELSE 0 END) AS fora
                         FROM documentos_recebidos
                         WHERE data_recepcao LIKE ?
                         GROUP BY endereçado_a""", (f"{prefixo}%",))
            por_dep = {r['dep']: r for r in c.fetchall()}

            result = []
            for dep in DEPARTAMENTOS_RECEBIDOS:
                r = por_dep.get(dep)
                total  = r['total']  if r else 0
                dentro = (r['dentro'] or 0) if r else 0
                fora   = (r['fora']   or 0) if r else 0
                taxa = round((dentro / total * 100) if total > 0 else 0, 1)
                result.append({'departamento': dep, 'total': total, 'dentro_prazo': dentro, 'fora_prazo': fora, 'taxa': taxa})
            return result

    def get_remetentes_frequentes(self, ano=None, mes=None):
        with self._connect() as conn:
            c = conn.cursor()
            ano = ano or str(date.today().year)
            if mes and mes != "0":
                prefixo = f"{ano}-{int(mes):02d}"
            else:
                prefixo = ano
            c.execute('''SELECT proveniencia, COUNT(*) as total
                         FROM documentos_recebidos
                         WHERE proveniencia IS NOT NULL AND proveniencia != ''
                         AND data_recepcao LIKE ?
                         GROUP BY proveniencia ORDER BY total DESC LIMIT 10''', (f"{prefixo}%",))
            return [dict(r) for r in c.fetchall()]

    def check_alertas(self):
        with self._connect() as conn:
            c = conn.cursor()
            today = date.today().isoformat()

            c.execute('''SELECT id, numero, assunto, data_recepcao, tecnico FROM documentos_recebidos
                         WHERE (data_resposta IS NULL OR data_resposta = '')
                         AND prazo_status NOT IN ('Arquivado', 'Arquivo', 'Fora do Prazo')''')
            pendentes = [dict(r) for r in c.fetchall()]

            from datetime import timedelta
            proximos_3 = (date.today() + timedelta(days=3)).isoformat()
            c.execute('''SELECT id, assunto, data_reuniao, hora_local, organizador FROM reunioes
                         WHERE data_reuniao >= ? AND data_reuniao <= ? ORDER BY data_reuniao''',
                      (today, proximos_3))
            agora = datetime.now()
            reunioes = []
            for r in c.fetchall():
                r = dict(r)
                _, fim = get_meeting_datetimes(r.get('data_reuniao', ''), r.get('hora_local', ''))
                if fim is not None and agora > fim:
                    continue
                reunioes.append(r)
        return {'docs_pendentes': pendentes, 'reunioes_proximas': reunioes}

    # ---- Reciclagem ----
    # Nomes legíveis das tabelas para apresentação na interface
    RECICLAGEM_TIPOS = {
        'documentos_recebidos': 'Doc. Recebido',
        'documentos_enviados':  'Doc. Enviado',
        'reunioes':             'Reunião',
        'contactos':            'Contacto',
    }

    def _mover_para_reciclagem(self, tabela, id, campos_descricao):
        """Em vez de apagar definitivamente, guarda o registo completo (JSON)
        na tabela reciclagem, de onde pode ser restaurado durante 30 dias."""
        with self._connect() as conn:
            c = conn.cursor()
            c.execute(f"SELECT * FROM {tabela} WHERE id=?", (id,))
            row = c.fetchone()
            if not row:
                return False
            d = dict(row)
            partes = [str(d.get(f) or '').strip() for f in campos_descricao]
            descricao = ' — '.join(p for p in partes if p) or f'registo {id}'
            c.execute("INSERT INTO reciclagem (tabela, registo_id, descricao, dados) "
                      "VALUES (?,?,?,?)",
                      (tabela, id, descricao[:120], json.dumps(d, ensure_ascii=False)))
            c.execute(f"DELETE FROM {tabela} WHERE id=?", (id,))
        return True

    def get_reciclagem(self):
        """Lista os itens na reciclagem, mais recentes primeiro."""
        with self._connect() as conn:
            c = conn.cursor()
            c.execute("SELECT id, tabela, registo_id, descricao, eliminado_em "
                      "FROM reciclagem ORDER BY eliminado_em DESC")
            return [dict(r) for r in c.fetchall()]

    def restaurar_reciclagem(self, rec_id):
        """Restaura um item da reciclagem para a tabela original. Devolve o
        nome legível do tipo restaurado, ou None se falhar."""
        with self._connect() as conn:
            c = conn.cursor()
            c.execute("SELECT tabela, dados FROM reciclagem WHERE id=?", (rec_id,))
            row = c.fetchone()
            if not row:
                return None
            tabela, dados = row['tabela'], json.loads(row['dados'])
            if tabela not in self.RECICLAGEM_TIPOS:
                return None
            # Usa apenas colunas que ainda existem no esquema actual
            c.execute(f"PRAGMA table_info({tabela})")
            cols_actuais = {r[1] for r in c.fetchall()}
            dados = {k: v for k, v in dados.items() if k in cols_actuais}
            # Mantém o id original salvo se, entretanto, tiver sido reutilizado
            if 'id' in dados:
                c.execute(f"SELECT 1 FROM {tabela} WHERE id=?", (dados['id'],))
                if c.fetchone():
                    dados.pop('id')
            cols = ', '.join(f'"{k}"' for k in dados)
            marks = ', '.join('?' for _ in dados)
            c.execute(f'INSERT INTO {tabela} ({cols}) VALUES ({marks})',
                      list(dados.values()))
            c.execute("DELETE FROM reciclagem WHERE id=?", (rec_id,))
            return self.RECICLAGEM_TIPOS[tabela]

    def eliminar_reciclagem(self, rec_id):
        """Remove definitivamente um item da reciclagem."""
        with self._connect() as conn:
            c = conn.cursor()
            c.execute("DELETE FROM reciclagem WHERE id=?", (rec_id,))
        return True

    def esvaziar_reciclagem(self):
        """Esvazia toda a reciclagem. Devolve o número de itens removidos."""
        with self._connect() as conn:
            c = conn.cursor()
            c.execute("SELECT COUNT(*) FROM reciclagem")
            n = c.fetchone()[0]
            c.execute("DELETE FROM reciclagem")
            return n

    # ---- Utilitários ----
    def suggest_next_numero(self, tabela='recebidos'):
        """Sugere o próximo número de documento com base nos já existentes."""
        import re as _re
        table = 'documentos_recebidos' if tabela == 'recebidos' else 'documentos_enviados'
        with self._connect() as conn:
            c = conn.cursor()
            c.execute(f"SELECT numero FROM {table} WHERE numero IS NOT NULL AND numero != '' ORDER BY id DESC LIMIT 100")
            numeros = [r[0] for r in c.fetchall()]
        if not numeros:
            return ""
        template = numeros[0]
        ano_atual = str(date.today().year)
        partes = _re.findall(r'\d+', template)
        seq_candidates = [p for p in partes if not (len(p) == 4 and p[:2] in ('19', '20'))]
        if not seq_candidates:
            return ""
        seq = seq_candidates[0]
        next_seq = str(int(seq) + 1).zfill(len(seq))
        replaced = [False]
        def _rep(m):
            part = m.group(0)
            # Actualiza o ano (4 dígitos começados por 19/20) para o ano actual
            if len(part) == 4 and part[:2] in ('19', '20'):
                return ano_atual
            # Incrementa a primeira parte sequencial (não-ano)
            if not replaced[0] and part == seq:
                replaced[0] = True
                return next_seq
            return part
        return _re.sub(r'\d+', _rep, template)

    def backup_para(self, destino):
        """Cria uma cópia de segurança consistente da base de dados usando a
        API de backup do SQLite (segura mesmo com a base de dados em uso,
        ao contrário de uma cópia simples do ficheiro)."""
        dest = sqlite3.connect(destino)
        try:
            src = sqlite3.connect(self.db_path)
            try:
                src.backup(dest)
            finally:
                src.close()
        finally:
            dest.close()

    def vacuum(self):
        """Executa VACUUM no SQLite para compactar e optimizar a base de dados."""
        conn = sqlite3.connect(self.db_path)
        try:
            conn.execute("VACUUM")
        finally:
            conn.close()

    def get_relatorio_evolucao_mensal(self, ano=None):
        """Devolve dados mensais (recebidos, respondidos, fora de prazo) para o ano indicado."""
        ano = ano or str(date.today().year)
        with self._connect() as conn:
            c = conn.cursor()
            # Uma única consulta agregada por mês (em vez de 36 consultas)
            c.execute("""SELECT substr(data_recepcao, 6, 2) AS mes,
                                COUNT(*) AS recebidos,
                                SUM(CASE WHEN data_resposta IS NOT NULL AND data_resposta != ''
                                         THEN 1 ELSE 0 END) AS respondidos,
                                SUM(CASE WHEN prazo_status='Fora do Prazo' THEN 1 ELSE 0 END) AS fora
                         FROM documentos_recebidos
                         WHERE data_recepcao LIKE ?
                         GROUP BY mes""", (f"{ano}-%",))
            por_mes = {}
            for r in c.fetchall():
                try:
                    por_mes[int(r['mes'])] = r
                except (TypeError, ValueError):
                    pass
            result = []
            for mes in range(1, 13):
                r = por_mes.get(mes)
                result.append({'mes': mes,
                               'recebidos':   r['recebidos']          if r else 0,
                               'respondidos': (r['respondidos'] or 0) if r else 0,
                               'fora_prazo':  (r['fora'] or 0)        if r else 0})
            return result

    def get_relatorio_tecnicos(self, ano=None, mes=None):
        """Devolve ranking de técnicos com total de documentos e taxa de cumprimento."""
        ano = ano or str(date.today().year)
        prefixo = f"{ano}-{int(mes):02d}" if mes and mes != "0" else ano
        with self._connect() as conn:
            c = conn.cursor()
            c.execute("""SELECT tecnico,
                         COUNT(*) as total,
                         SUM(CASE WHEN prazo_status IN ('Dentro do Prazo', 'Arquivado', 'Arquivo') THEN 1 ELSE 0 END) as dentro,
                         SUM(CASE WHEN prazo_status='Fora do Prazo' THEN 1 ELSE 0 END) as fora
                         FROM documentos_recebidos
                         WHERE tecnico IS NOT NULL AND tecnico != ''
                         AND data_recepcao LIKE ?
                         GROUP BY tecnico ORDER BY total DESC""", (f"{prefixo}%",))
            rows = []
            for r in c.fetchall():
                dentro, fora = r['dentro'], r['fora']
                taxa = round((dentro / (dentro + fora) * 100) if (dentro + fora) > 0 else 0, 1)
                rows.append({'tecnico': r['tecnico'], 'total': r['total'],
                             'dentro': dentro, 'fora': fora, 'taxa': taxa})
            return rows

    def export_all_excel(self, filepath):
        """Exporta todos os dados (Recebidos, Enviados, Reuniões, Contactos) para um Excel com 4 folhas."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            hf = Font(bold=True, color="FFFFFF")
            hfill = PatternFill("solid", fgColor="1F4E79")
            ha = Alignment(horizontal='center')

            def _header(ws, cols):
                ws.append(cols)
                for cell in ws[1]:
                    cell.font = hf; cell.fill = hfill; cell.alignment = ha

            wb = openpyxl.Workbook()

            # Recebidos
            ws1 = wb.active; ws1.title = "Recebidos"
            _header(ws1, ['ID','Nº Doc','Proveniência','Remetente','Cargo','Assunto',
                          'Data Recepção','Despacho','Departamento','Técnico',
                          'Data Resposta','Status','Observação'])
            for r in self.get_all_recebidos():
                ws1.append([r.get('id'), r.get('numero'), r.get('proveniencia'),
                            r.get('remetente_nome'), r.get('remetente_cargo'), r.get('assunto'),
                            iso_to_display(r.get('data_recepcao')), r.get('despacho'), r.get('endereçado_a'),
                            r.get('tecnico'), iso_to_display(r.get('data_resposta')), r.get('prazo_status'),
                            r.get('observacao')])

            # Enviados
            ws2 = wb.create_sheet("Enviados")
            _header(ws2, ['ID','Nº Doc','Assunto','Preparado Por','Assinante',
                          'Destinatário','Cargo Dest.','Instituição','Data Envio','Observação'])
            for r in self.get_all_enviados():
                ws2.append([r.get('id'), r.get('numero'), r.get('assunto'),
                            r.get('preparado_por'), r.get('assinante'),
                            r.get('destinatario_nome'), r.get('destinatario_cargo'),
                            r.get('instituicao'), iso_to_display(r.get('data_envio')), r.get('observacao')])

            # Reuniões
            ws3 = wb.create_sheet("Reuniões")
            _header(ws3, ['ID','Nº Doc','Organizador','Data Conv.','Assunto',
                          'Data Reunião','Hora/Local','Participantes','Decisões','Cancelada'])
            for r in self.get_all_reunioes():
                ws3.append([r.get('id'), r.get('num_doc'), r.get('organizador'),
                            iso_to_display(r.get('data_convocatoria')), r.get('assunto'),
                            iso_to_display(r.get('data_reuniao')),
                            r.get('hora_local'), r.get('participantes'),
                            r.get('decisoes'), 'Sim' if r.get('cancelada') else 'Não'])

            # Contactos
            ws4 = wb.create_sheet("Contactos")
            _header(ws4, ['ID','Nº','Nome','Email','Telefone','Departamento','Cargo'])
            for r in self.get_all_contactos():
                ws4.append([r.get('id'), r.get('numero'), r.get('nome'),
                            r.get('email'), r.get('telefone'),
                            r.get('departamento'), r.get('cargo')])

            for ws in [ws1, ws2, ws3, ws4]:
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 18

            wb.save(filepath)
            return True
        except Exception:
            return False

    # ---- Exports ----
    def export_recebidos_excel(self, filepath, filters=None):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            rows = self.get_all_recebidos(filters)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Documentos Recebidos"
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="1F4E79")
            headers = ['ID', 'Nº Documento', 'Proveniência', 'Remetente', 'Cargo', 'Assunto',
                       'Data Recepção', 'Despacho', 'Ao Departamento', 'Técnico', 'Data Resposta',
                       'Status Prazo', 'Observação']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            for r in rows:
                ws.append([r.get('id'), r.get('numero'), r.get('proveniencia'), r.get('remetente_nome'),
                            r.get('remetente_cargo'), r.get('assunto'), iso_to_display(r.get('data_recepcao')),
                            r.get('despacho'), r.get('endereçado_a'), r.get('tecnico'),
                            iso_to_display(r.get('data_resposta')), r.get('prazo_status'), r.get('observacao')])
            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
            wb.save(filepath)
            return True
        except Exception as e:
            print(f"Erro ao exportar: {e}")
            return False

    def export_enviados_excel(self, filepath, filters=None):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            rows = self.get_all_enviados(filters)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Documentos Enviados"
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="1F4E79")
            headers = ['ID', 'Nº Doc', 'Assunto', 'Preparado Por', 'Assinante',
                       'Destinatário', 'Cargo Dest.', 'Instituição', 'Data Envio', 'Observação']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            for r in rows:
                ws.append([r.get('id'), r.get('numero'), r.get('assunto'), r.get('preparado_por'),
                            r.get('assinante'), r.get('destinatario_nome'), r.get('destinatario_cargo'),
                            r.get('instituicao'), iso_to_display(r.get('data_envio')), r.get('observacao')])
            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
            wb.save(filepath)
            return True
        except Exception as e:
            print(f"Erro ao exportar: {e}")
            return False

    def export_contactos_excel(self, filepath):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            rows = self.get_all_contactos()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Contactos"
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="1F4E79")
            headers = ['Nº', 'Nome', 'Email', 'Telefone', 'Departamento', 'Cargo']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            for r in rows:
                ws.append([r.get('numero'), r.get('nome'), r.get('email'),
                            r.get('telefone'), r.get('departamento'), r.get('cargo')])
            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
            wb.save(filepath)
            return True
        except Exception as e:
            print(f"Erro ao exportar: {e}")
            return False
