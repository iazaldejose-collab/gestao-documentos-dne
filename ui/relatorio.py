import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import date
import customtkinter as ctk

from ui.widgets import BusyDialog

# matplotlib é pesado e demora a importar — só é carregado quando o
# utilizador abre a página de Relatório (ver _build_chart)
plt = None
FigureCanvasTkAgg = None
HAS_MATPLOTLIB = None
_MATPLOTLIB_ERRO = ""


def _carregar_matplotlib():
    global plt, FigureCanvasTkAgg, HAS_MATPLOTLIB, _MATPLOTLIB_ERRO
    if HAS_MATPLOTLIB is None:
        try:
            import os
            os.environ.setdefault("MPLBACKEND", "TkAgg")
            import matplotlib
            matplotlib.use("TkAgg")
            import matplotlib.pyplot as _plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg as _FigureCanvasTkAgg
            plt = _plt
            FigureCanvasTkAgg = _FigureCanvasTkAgg
            HAS_MATPLOTLIB = True
        except Exception as e:
            HAS_MATPLOTLIB = False
            _MATPLOTLIB_ERRO = str(e)
    return HAS_MATPLOTLIB

MESES = [
    ("Todos os Meses", "0"),
    ("Janeiro",   "1"),  ("Fevereiro",  "2"),  ("Março",     "3"),
    ("Abril",     "4"),  ("Maio",       "5"),  ("Junho",     "6"),
    ("Julho",     "7"),  ("Agosto",     "8"),  ("Setembro",  "9"),
    ("Outubro",  "10"),  ("Novembro",  "11"),  ("Dezembro", "12"),
]
MESES_LABELS = [m[0] for m in MESES]
MESES_MAP    = {m[0]: m[1] for m in MESES}


class RelatorioFrame(ctk.CTkFrame):
    def __init__(self, parent, db, config):
        super().__init__(parent, corner_radius=0)
        self.db = db
        self.config = config
        self._figs = {}    # figuras matplotlib activas, por chave de gráfico
        self._tipos = {}   # tipo de gráfico escolhido, por chave (persiste no refresh)

        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self._build_toolbar()
        self._build_body()

    # ------------------------------------------------------------------ toolbar
    def _build_toolbar(self):
        tb = ctk.CTkFrame(self, height=56, corner_radius=0, fg_color=("gray90", "gray20"))
        tb.grid(row=0, column=0, sticky="ew")
        tb.grid_columnconfigure(4, weight=1)

        ctk.CTkLabel(tb, text="📊  Relatório e Dashboard",
                     font=ctk.CTkFont(size=15, weight="bold")).grid(
                         row=0, column=0, padx=15, pady=10, sticky="w")

        # --- Ano ---
        ctk.CTkLabel(tb, text="Ano:", font=ctk.CTkFont(size=12)).grid(
            row=0, column=1, padx=(20, 4), pady=10)

        anos = self.db.get_anos_disponiveis()
        self.ano_var = tk.StringVar(value=str(date.today().year))
        self.cmb_ano = ctk.CTkComboBox(tb, values=anos, variable=self.ano_var,
                                       width=90, command=lambda e: self.refresh())
        self.cmb_ano.grid(row=0, column=2, padx=4, pady=10)

        # --- Mês ---
        ctk.CTkLabel(tb, text="Mês:", font=ctk.CTkFont(size=12)).grid(
            row=0, column=3, padx=(12, 4), pady=10)

        self.mes_var = tk.StringVar(value="Todos os Meses")
        self.cmb_mes = ctk.CTkComboBox(tb, values=MESES_LABELS, variable=self.mes_var,
                                       width=160, command=lambda e: self.refresh())
        self.cmb_mes.grid(row=0, column=4, padx=4, pady=10, sticky="w")

        # --- botões ---
        btn_frame = ctk.CTkFrame(tb, fg_color="transparent")
        btn_frame.grid(row=0, column=5, padx=10, pady=6, sticky="e")
        ctk.CTkButton(btn_frame, text="🔄 Actualizar", width=110, command=self.refresh,
                      fg_color="#1F4E79").pack(side="left", padx=4)
        ctk.CTkButton(btn_frame, text="📤 Exportar Excel", width=130, command=self.exportar,
                      fg_color="#27ae60").pack(side="left", padx=4)
        ctk.CTkButton(btn_frame, text="📄 Exportar PDF", width=120, command=self.exportar_pdf,
                      fg_color="#c0392b").pack(side="left", padx=4)

    # ------------------------------------------------------------------ body
    def _build_body(self):
        self.scroll = ctk.CTkScrollableFrame(self, corner_radius=0)
        self.scroll.grid(row=1, column=0, sticky="nsew")
        self.scroll.grid_columnconfigure(0, weight=1)
        self.refresh()

    # ------------------------------------------------------------------ helpers
    def _get_filtros(self):
        ano = self.ano_var.get().strip()
        mes = MESES_MAP.get(self.mes_var.get(), "0")
        return ano, mes

    def _periodo_label(self):
        """Texto descritivo do período seleccionado, ex: 'Ano 2026' ou 'Março 2026'."""
        ano, mes = self._get_filtros()
        if mes == "0":
            return f"Ano {ano}"
        nome_mes = self.mes_var.get()
        return f"{nome_mes} de {ano}"

    # ------------------------------------------------------------------ refresh
    def refresh(self, *args):
        for w in self.scroll.winfo_children():
            w.destroy()
        # Liberta todas as figuras matplotlib do refresh anterior (sem isto,
        # cada actualização do Relatório acumulava figuras em memória)
        for fig in list(self._figs.values()):
            if fig is not None:
                try:
                    plt.close(fig)
                except Exception:
                    pass
        self._figs.clear()

        try:
            self._build_kpis()
            self._build_dept_table()
            if _carregar_matplotlib():
                self._build_chart_dept()
                self._build_chart_taxa()
                self._build_chart()
                self._build_chart_evolucao()
            elif _MATPLOTLIB_ERRO:
                ctk.CTkLabel(self.scroll,
                             text=f"⚠️ Gráficos indisponíveis: {_MATPLOTLIB_ERRO}",
                             text_color="orange",
                             font=ctk.CTkFont(size=11)).pack(pady=4)
            self._build_tecnico_table()
            self._build_remetentes()
        except Exception as e:
            ctk.CTkLabel(self.scroll, text=f"Erro ao carregar relatório: {e}",
                         text_color="red").pack(pady=20)

    # ------------------------------------------------------------------ KPIs
    def _build_kpis(self):
        ano, mes = self._get_filtros()
        stats = self.db.get_relatorio_stats(ano=ano, mes=mes)
        periodo = self._periodo_label()

        section = ctk.CTkFrame(self.scroll, corner_radius=8)
        section.pack(fill="x", padx=15, pady=(15, 8))

        ctk.CTkLabel(section,
                     text=f"Indicadores Chave de Desempenho  —  {periodo}",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(
                         anchor="w", padx=10, pady=(8, 4))

        cards_frame = ctk.CTkFrame(section, fg_color="transparent")
        cards_frame.pack(fill="x", padx=10, pady=(0, 10))

        label_rec = f"📥 Recebidos\n({periodo})"
        label_env = f"📤 Enviados\n({periodo})"
        label_reu = f"📅 Reuniões\n({periodo})"

        kpis = [
            (label_rec,          str(stats['total_recebidos']),   "#1F4E79"),
            ("✅ Respondidos",   str(stats['total_respondidos']),  "#27ae60"),
            ("📈 Taxa Cumprimento",
             f"{stats['taxa_cumprimento']}%",
             "#27ae60" if stats['taxa_cumprimento'] >= 80 else "#e67e22"),
            ("⚠️ Fora do Prazo", str(stats['total_fora_prazo']),  "#c0392b"),
            (label_reu,          str(stats['reunioes_mes']),       "#8e44ad"),
            (label_env,          str(stats['total_enviados']),     "#2c6fad"),
        ]

        for label, value, color in kpis:
            card = ctk.CTkFrame(cards_frame, width=150, height=90, corner_radius=10,
                                fg_color=color)
            card.pack(side="left", padx=6, pady=4)
            card.pack_propagate(False)
            ctk.CTkLabel(card, text=value, font=ctk.CTkFont(size=26, weight="bold"),
                         text_color="white").pack(expand=True)
            ctk.CTkLabel(card, text=label, font=ctk.CTkFont(size=10),
                         text_color="white", wraplength=140, justify="center").pack(pady=(0, 6))

    # ------------------------------------------------------------------ tabela depts
    def _build_dept_table(self):
        ano, mes = self._get_filtros()
        depts = self.db.get_relatorio_departamentos(ano=ano, mes=mes)
        periodo = self._periodo_label()

        section = ctk.CTkFrame(self.scroll, corner_radius=8)
        section.pack(fill="x", padx=15, pady=8)

        ctk.CTkLabel(section, text=f"Desempenho por Departamento  —  {periodo}",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(
                         anchor="w", padx=10, pady=(8, 4))

        table_frame = ctk.CTkFrame(section, fg_color="transparent")
        table_frame.pack(fill="x", padx=10, pady=(0, 10))

        style = ttk.Style()
        style.configure("Dept.Treeview", rowheight=24, font=('Segoe UI', 10))
        style.configure("Dept.Treeview.Heading", font=('Segoe UI', 10, 'bold'),
                        background="#1F4E79", foreground="white")

        cols = ("dept", "total", "dentro", "fora", "taxa")
        tree = ttk.Treeview(table_frame, columns=cols, show="headings",
                            style="Dept.Treeview", height=6)
        col_cfg = [("dept", "Departamento", 260), ("total", "Total", 70),
                   ("dentro", "Dentro Prazo", 100), ("fora", "Fora Prazo", 100),
                   ("taxa", "Taxa (%)", 80)]
        for col, heading, width in col_cfg:
            tree.heading(col, text=heading, anchor="center")
            anchor = "w" if col == "dept" else "center"
            tree.column(col, width=width, minwidth=40, anchor=anchor)

        tree.tag_configure("bom", foreground="#1d8348")
        tree.tag_configure("medio", foreground="#cb6e1c")
        tree.tag_configure("fraco", foreground="#c0392b")
        tree.tag_configure("vazio", foreground="gray")

        for d in depts:
            if d['total'] == 0:
                tag, taxa_txt = "vazio", "—"
            elif d['taxa'] >= 80:
                tag, taxa_txt = "bom", f"{d['taxa']}%"
            elif d['taxa'] >= 50:
                tag, taxa_txt = "medio", f"{d['taxa']}%"
            else:
                tag, taxa_txt = "fraco", f"{d['taxa']}%"
            tree.insert("", "end", values=(
                d['departamento'], d['total'], d['dentro_prazo'],
                d['fora_prazo'], taxa_txt
            ), tags=(tag,))
        tree.pack(fill="x")

        legenda = ctk.CTkFrame(section, fg_color="transparent")
        legenda.pack(fill="x", padx=10, pady=(0, 8))
        for texto, cor in (("● ≥ 80% — Bom", "#1d8348"),
                           ("● 50–79% — Atenção", "#cb6e1c"),
                           ("● < 50% — Crítico", "#c0392b")):
            ctk.CTkLabel(legenda, text=texto, font=ctk.CTkFont(size=10),
                         text_color=cor).pack(side="left", padx=(0, 14))

    # ================= Gráficos com selector de tipo =================
    TIPOS_CONTAGEM = ["Colunas Agrupadas", "Colunas Empilhadas", "Barras Horizontais",
                      "Linha", "Área", "Pizza (Volume)", "Rosca (Volume)", "Radar (Taxa)"]
    TIPOS_TAXA = ["Barras Horizontais", "Colunas", "Linha", "Área"]
    TIPOS_EVOL = ["Linha", "Área", "Colunas Agrupadas", "Colunas Empilhadas", "Barras Horizontais"]

    @staticmethod
    def _nome_dept(nome):
        """Nome do departamento limpo para os gráficos: remove os prefixos
        'Dep.', 'Rep.', 'Direcção -' e a preposição inicial 'de'
        (ex.: 'Dep. de Planeamento Energético' -> 'Planeamento Energético')."""
        import re
        n = (nome or '').strip()
        n = re.sub(r'^(Dep\.?|Rep\.?|Direc[çc][ãa]o)\s*[-–]?\s*', '', n, flags=re.IGNORECASE)
        n = re.sub(r'^de\s+', '', n, flags=re.IGNORECASE)
        return n.strip() or (nome or '')

    def _secao_grafico(self, titulo, tipos, key, render):
        """Cria uma secção de gráfico com selector de tipo (ComboBox) e redesenha
        ao mudar de tipo. 'render(tipo)' devolve uma figura matplotlib."""
        if key not in self._tipos:
            self._tipos[key] = tk.StringVar(value=tipos[0])
        elif self._tipos[key].get() not in tipos:
            self._tipos[key].set(tipos[0])
        section = ctk.CTkFrame(self.scroll, corner_radius=8)
        section.pack(fill="x", padx=15, pady=8)
        header = ctk.CTkFrame(section, fg_color="transparent")
        header.pack(fill="x", padx=10, pady=(8, 6))
        ctk.CTkLabel(header, text=titulo,
                     font=ctk.CTkFont(size=13, weight="bold")).pack(side="left")
        ctk.CTkLabel(header, text="Tipo:", font=ctk.CTkFont(size=11)).pack(side="right", padx=(6, 4))
        frame = ctk.CTkFrame(section, fg_color="transparent")

        def _draw(*_):
            try:
                if not frame.winfo_exists():
                    return
            except Exception:
                return
            old = self._figs.get(key)
            if old is not None:
                try:
                    plt.close(old)
                except Exception:
                    pass
                self._figs[key] = None
            for w in frame.winfo_children():
                w.destroy()
            try:
                fig = render(self._tipos[key].get())
                self._figs[key] = fig
                canvas = FigureCanvasTkAgg(fig, master=frame)
                canvas.get_tk_widget().pack(fill="x")
                canvas.draw()
            except Exception as e:
                ctk.CTkLabel(frame, text=f"Gráfico indisponível: {e}",
                             text_color="red").pack(pady=10)

        ctk.CTkComboBox(header, values=tipos, variable=self._tipos[key], width=200,
                        command=lambda v: _draw()).pack(side="right")
        frame.pack(fill="x", padx=10, pady=(0, 10))
        _draw()

    def _build_chart_dept(self):
        ano, mes = self._get_filtros()
        depts = [d for d in self.db.get_relatorio_departamentos(ano=ano, mes=mes) if d['total'] > 0]
        if not depts:
            return
        periodo = self._periodo_label()
        self._secao_grafico(f"Gráfico de Desempenho por Departamento  —  {periodo}",
                            self.TIPOS_CONTAGEM, 'dept',
                            lambda tipo: self._fig_contagem(tipo, depts, periodo,
                                                            "Documentos por Departamento"))

    def _fig_contagem(self, tipo, depts, periodo, titulo):
        """Figura para dados por departamento (Dentro/Fora/Volume/Taxa)."""
        import math
        labels = [self._nome_dept(d['departamento']) for d in depts]
        dentro = [d['dentro_prazo'] for d in depts]
        fora   = [d['fora_prazo']   for d in depts]
        taxa   = [d['taxa']         for d in depts]
        total  = [d['total']        for d in depts]
        cores_taxa = ['#27ae60' if t >= 80 else '#cb6e1c' if t >= 50 else '#c0392b' for t in taxa]
        x = list(range(len(labels)))

        if tipo == "Colunas Empilhadas":
            fig, ax = plt.subplots(figsize=(9, 3.5))
            ax.bar(x, dentro, label='Dentro do Prazo', color='#27ae60')
            ax.bar(x, fora, bottom=dentro, label='Fora do Prazo', color='#c0392b')
            ax.set_xticks(x); ax.set_xticklabels(labels, rotation=20, ha='right', fontsize=8)
            ax.set_ylabel('Documentos'); ax.set_title(f'{titulo} — {periodo}'); ax.legend()
        elif tipo == "Barras Horizontais":
            fig, ax = plt.subplots(figsize=(9, max(2.8, 0.5 * len(labels) + 1)))
            w = 0.4
            ax.barh([i + w/2 for i in x], dentro, w, label='Dentro do Prazo', color='#27ae60')
            ax.barh([i - w/2 for i in x], fora,   w, label='Fora do Prazo',   color='#c0392b')
            ax.set_yticks(x); ax.set_yticklabels(labels, fontsize=8); ax.invert_yaxis()
            ax.set_xlabel('Documentos'); ax.set_title(f'{titulo} — {periodo}'); ax.legend()
        elif tipo == "Linha":
            fig, ax = plt.subplots(figsize=(9, 3.5))
            ax.plot(x, dentro, 'o-', color='#27ae60', label='Dentro do Prazo', linewidth=2)
            ax.plot(x, fora, 's--', color='#c0392b', label='Fora do Prazo', linewidth=2)
            ax.set_xticks(x); ax.set_xticklabels(labels, rotation=20, ha='right', fontsize=8)
            ax.set_ylabel('Documentos'); ax.set_title(f'{titulo} — {periodo}')
            ax.legend(); ax.grid(axis='y', alpha=0.3)
        elif tipo == "Área":
            fig, ax = plt.subplots(figsize=(9, 3.5))
            ax.stackplot(x, dentro, fora, labels=['Dentro do Prazo', 'Fora do Prazo'],
                         colors=['#27ae60', '#c0392b'], alpha=0.8)
            ax.set_xticks(x); ax.set_xticklabels(labels, rotation=20, ha='right', fontsize=8)
            ax.set_ylabel('Documentos'); ax.set_title(f'{titulo} — {periodo}'); ax.legend()
        elif tipo in ("Pizza (Volume)", "Rosca (Volume)"):
            fig, ax = plt.subplots(figsize=(7, 5))
            kw = dict(wedgeprops=dict(width=0.5), pctdistance=0.80) if "Rosca" in tipo else {}
            wedges, texts, autotexts = ax.pie(total, labels=labels, autopct='%1.1f%%',
                                              colors=cores_taxa, startangle=90, **kw)
            for t in texts:
                t.set_fontsize(7)
            for at in autotexts:
                at.set_fontsize(7)
            if "Rosca" in tipo:
                ax.text(0, 0, f"Total\n{sum(total)}", ha='center', va='center',
                        fontsize=12, fontweight='bold')
            ax.set_title(f'Volume por Departamento — {periodo}')
        elif tipo == "Radar (Taxa)":
            N = len(labels)
            if N < 3:
                raise ValueError("Radar requer pelo menos 3 departamentos com dados.")
            angles = [n / float(N) * 2 * math.pi for n in range(N)]
            angles += angles[:1]
            tp = taxa + taxa[:1]
            fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
            ax.plot(angles, tp, 'o-', linewidth=2, color='#2c6fad')
            ax.fill(angles, tp, alpha=0.25, color='#2c6fad')
            ax.set_xticks(angles[:-1]); ax.set_xticklabels(labels, size=7)
            ax.set_ylim(0, 100); ax.set_yticks([20, 40, 60, 80, 100])
            ax.set_yticklabels(['20%', '40%', '60%', '80%', '100%'], fontsize=7)
            ax.set_title(f'Perfil de Desempenho — {periodo}', pad=20)
        else:  # "Colunas Agrupadas" (padrão)
            fig, ax = plt.subplots(figsize=(9, 3.5))
            w = 0.35
            ax.bar([i - w/2 for i in x], dentro, w, label='Dentro do Prazo', color='#27ae60')
            ax.bar([i + w/2 for i in x], fora,   w, label='Fora do Prazo',   color='#c0392b')
            ax.set_xticks(x); ax.set_xticklabels(labels, rotation=20, ha='right', fontsize=8)
            ax.set_ylabel('Documentos'); ax.set_title(f'{titulo} — {periodo}'); ax.legend()
        fig.tight_layout()
        return fig

    # ------------------------------------------------------------------ gráfico taxa
    def _build_chart_taxa(self):
        ano, mes = self._get_filtros()
        depts = [d for d in self.db.get_relatorio_departamentos(ano=ano, mes=mes) if d['total'] > 0]
        if not depts:
            return
        periodo = self._periodo_label()
        ordenados = sorted(depts, key=lambda d: d['taxa'], reverse=True)
        self._secao_grafico(f"Ranking — Taxa de Cumprimento por Departamento  —  {periodo}",
                            self.TIPOS_TAXA, 'taxa',
                            lambda tipo: self._fig_taxa(tipo, ordenados, periodo))

    def _fig_taxa(self, tipo, depts, periodo):
        """Figura para o ranking de taxa de cumprimento (série única: taxa %)."""
        labels = [self._nome_dept(d['departamento']) for d in depts]
        taxa = [d['taxa'] for d in depts]
        cores = ['#27ae60' if t >= 80 else '#cb6e1c' if t >= 50 else '#c0392b' for t in taxa]
        x = list(range(len(labels)))

        if tipo == "Colunas":
            fig, ax = plt.subplots(figsize=(9, 3.5))
            ax.bar(x, taxa, color=cores)
            ax.set_xticks(x); ax.set_xticklabels(labels, rotation=20, ha='right', fontsize=8)
            ax.set_ylim(0, 100); ax.set_ylabel('Taxa de Cumprimento (%)')
            ax.axhline(80, color='gray', linestyle='--', linewidth=1)
            for i, t in enumerate(taxa):
                ax.text(i, min(t, 100) + 1, f"{t}%", ha='center', fontsize=8)
            ax.set_title(f'Ranking de Cumprimento — {periodo}')
        elif tipo == "Linha":
            fig, ax = plt.subplots(figsize=(9, 3.5))
            ax.plot(x, taxa, 'o-', color='#2c6fad', linewidth=2)
            ax.set_xticks(x); ax.set_xticklabels(labels, rotation=20, ha='right', fontsize=8)
            ax.set_ylim(0, 100); ax.set_ylabel('Taxa (%)')
            ax.axhline(80, color='gray', linestyle='--', linewidth=1)
            ax.grid(axis='y', alpha=0.3)
            ax.set_title(f'Ranking de Cumprimento — {periodo}')
        elif tipo == "Área":
            fig, ax = plt.subplots(figsize=(9, 3.5))
            ax.fill_between(x, taxa, color='#2c6fad', alpha=0.3)
            ax.plot(x, taxa, 'o-', color='#2c6fad', linewidth=2)
            ax.set_xticks(x); ax.set_xticklabels(labels, rotation=20, ha='right', fontsize=8)
            ax.set_ylim(0, 100); ax.set_ylabel('Taxa (%)')
            ax.axhline(80, color='gray', linestyle='--', linewidth=1)
            ax.set_title(f'Ranking de Cumprimento — {periodo}')
        else:  # "Barras Horizontais" (padrão)
            fig, ax = plt.subplots(figsize=(9, max(2.5, 0.45 * len(labels) + 1)))
            ax.barh(x, taxa, color=cores)
            ax.set_yticks(x); ax.set_yticklabels(labels, fontsize=8); ax.invert_yaxis()
            ax.set_xlim(0, 100); ax.set_xlabel('Taxa de Cumprimento (%)')
            ax.axvline(80, color='gray', linestyle='--', linewidth=1)
            ax.text(80, -0.7, 'Meta: 80%', ha='center', fontsize=8, color='gray')
            for i, t in enumerate(taxa):
                ax.text(min(t, 100) + 1, i, f"{t}%", va='center', fontsize=8)
            ax.set_title(f'Ranking de Cumprimento — {periodo}')
        fig.tight_layout()
        return fig

    # ------------------------------------------------------------------ gráfico
    def _build_chart(self):
        ano, mes = self._get_filtros()
        depts = [d for d in self.db.get_relatorio_departamentos(ano=ano, mes=mes) if d['total'] > 0]
        if not depts:
            return
        periodo = self._periodo_label()
        self._secao_grafico(f"Gráfico: Dentro vs Fora do Prazo  —  {periodo}",
                            self.TIPOS_CONTAGEM, 'prazo',
                            lambda tipo: self._fig_contagem(tipo, depts, periodo,
                                                            "Cumprimento por Departamento"))

    # ------------------------------------------------------------------ técnicos (B3)
    def _build_tecnico_table(self):
        ano, mes = self._get_filtros()
        tecnicos = self.db.get_relatorio_tecnicos(ano=ano, mes=mes)
        if not tecnicos:
            return
        periodo = self._periodo_label()
        section = ctk.CTkFrame(self.scroll, corner_radius=8)
        section.pack(fill="x", padx=15, pady=8)
        ctk.CTkLabel(section, text=f"Desempenho por Técnico  —  {periodo}",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(anchor="w", padx=10, pady=(8, 4))

        style = ttk.Style()
        style.configure("Tec.Treeview", rowheight=24, font=('Segoe UI', 10))
        style.configure("Tec.Treeview.Heading", font=('Segoe UI', 10, 'bold'),
                        background="#1F4E79", foreground="white")
        cols = ("tec", "total", "dentro", "fora", "taxa")
        tree = ttk.Treeview(section, columns=cols, show="headings",
                            style="Tec.Treeview", height=min(len(tecnicos), 6))
        for col, heading, width in [("tec","Técnico",200),("total","Total",70),
                                     ("dentro","Dentro Prazo",100),("fora","Fora Prazo",100),
                                     ("taxa","Taxa (%)",80)]:
            tree.heading(col, text=heading, anchor="center")
            tree.column(col, width=width, anchor="w" if col == "tec" else "center")
        tree.tag_configure("bom", foreground="#1d8348")
        tree.tag_configure("medio", foreground="#cb6e1c")
        tree.tag_configure("fraco", foreground="#c0392b")
        for t in tecnicos:
            tag = "bom" if t['taxa'] >= 80 else ("medio" if t['taxa'] >= 50 else "fraco")
            taxa_txt = f"{t['taxa']}%" if (t['dentro'] + t['fora']) > 0 else "—"
            tree.insert("", "end", values=(t['tecnico'], t['total'], t['dentro'], t['fora'], taxa_txt), tags=(tag,))
        tree.pack(fill="x", padx=10, pady=(0, 10))

    # ------------------------------------------------------------------ evolução mensal (B1)
    def _build_chart_evolucao(self):
        ano, mes = self._get_filtros()
        dados = self.db.get_relatorio_evolucao_mensal(ano=ano)
        if not any(d['recebidos'] > 0 for d in dados):
            return
        self._secao_grafico(f"Evolução Mensal  —  Ano {ano}",
                            self.TIPOS_EVOL, 'evol',
                            lambda tipo: self._fig_evol(tipo, dados, ano))

    def _fig_evol(self, tipo, dados, ano):
        """Figura da evolução mensal (Recebidos / Respondidos / Fora do Prazo)."""
        MESES_ABREV = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
                       'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        xs = list(range(12))
        rec  = [d['recebidos']   for d in dados]
        resp = [d['respondidos'] for d in dados]
        fora = [d['fora_prazo']  for d in dados]

        if tipo == "Área":
            fig, ax = plt.subplots(figsize=(9, 3.2))
            ax.stackplot(xs, rec, resp, fora,
                         labels=['Recebidos', 'Respondidos', 'Fora do Prazo'],
                         colors=['#2c6fad', '#27ae60', '#c0392b'], alpha=0.7)
        elif tipo == "Colunas Agrupadas":
            fig, ax = plt.subplots(figsize=(9, 3.2))
            w = 0.27
            ax.bar([i - w for i in xs], rec,  w, label='Recebidos',    color='#2c6fad')
            ax.bar(xs,                  resp, w, label='Respondidos',   color='#27ae60')
            ax.bar([i + w for i in xs], fora, w, label='Fora do Prazo', color='#c0392b')
        elif tipo == "Colunas Empilhadas":
            fig, ax = plt.subplots(figsize=(9, 3.2))
            ax.bar(xs, resp, label='Respondidos', color='#27ae60')
            ax.bar(xs, fora, bottom=resp, label='Fora do Prazo', color='#c0392b')
            ax.plot(xs, rec, 'o-', color='#2c6fad', label='Recebidos', linewidth=2)
        elif tipo == "Barras Horizontais":
            fig, ax = plt.subplots(figsize=(9, 4.6))
            w = 0.27
            ax.barh([i - w for i in xs], rec,  w, label='Recebidos',    color='#2c6fad')
            ax.barh(xs,                  resp, w, label='Respondidos',   color='#27ae60')
            ax.barh([i + w for i in xs], fora, w, label='Fora do Prazo', color='#c0392b')
            ax.set_yticks(xs); ax.set_yticklabels(MESES_ABREV, fontsize=8); ax.invert_yaxis()
            ax.set_xlabel('Documentos'); ax.set_title(f'Evolução Mensal — {ano}')
            ax.legend(fontsize=8)
            fig.tight_layout()
            return fig
        else:  # "Linha" (padrão)
            fig, ax = plt.subplots(figsize=(9, 3.2))
            ax.plot(xs, rec,  'o-',  color='#2c6fad', label='Recebidos',    linewidth=2)
            ax.plot(xs, resp, 's-',  color='#27ae60', label='Respondidos',   linewidth=2)
            ax.plot(xs, fora, '^--', color='#c0392b', label='Fora do Prazo', linewidth=1.5)
        ax.set_xticks(xs); ax.set_xticklabels(MESES_ABREV, fontsize=8)
        ax.set_ylabel('Documentos'); ax.set_title(f'Evolução Mensal — {ano}')
        ax.legend(fontsize=8); ax.grid(axis='y', alpha=0.3)
        fig.tight_layout()
        return fig

    # ------------------------------------------------------------------ remetentes
    def _build_remetentes(self):
        ano, mes = self._get_filtros()
        remetentes = self.db.get_remetentes_frequentes(ano=ano, mes=mes)
        periodo = self._periodo_label()

        section = ctk.CTkFrame(self.scroll, corner_radius=8)
        section.pack(fill="x", padx=15, pady=(8, 15))
        ctk.CTkLabel(section, text=f"Principais Remetentes  —  {periodo}",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(
                         anchor="w", padx=10, pady=(8, 4))

        style = ttk.Style()
        style.configure("Rem.Treeview", rowheight=24, font=('Segoe UI', 10))
        style.configure("Rem.Treeview.Heading", font=('Segoe UI', 10, 'bold'),
                        background="#1F4E79", foreground="white")

        cols = ("proveniencia", "total")
        height = max(len(remetentes), 1)
        tree = ttk.Treeview(section, columns=cols, show="headings",
                            style="Rem.Treeview", height=min(height, 8))
        tree.heading("proveniencia", text="Proveniência / Instituição", anchor="center")
        tree.column("proveniencia", width=300, anchor="w")
        tree.heading("total", text="Total Documentos", anchor="center")
        tree.column("total", width=130, anchor="center")
        for r in remetentes:
            tree.insert("", "end", values=(r['proveniencia'], r['total']))
        if not remetentes:
            tree.insert("", "end", values=("— Sem dados para o período —", ""))
        tree.pack(fill="x", padx=10, pady=(0, 10))

    # ------------------------------------------------------------------ exportar
    def exportar(self):
        ano, mes = self._get_filtros()
        periodo = self._periodo_label()
        nome_ficheiro = f"relatorio_dne_{periodo.replace(' ', '_').replace('/', '-')}.xlsx"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=nome_ficheiro,
            parent=self
        )
        if not filepath:
            return
        busy = BusyDialog(self, "A gerar relatório...")
        try:
            import openpyxl
            from openpyxl.styles import Font
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Relatório"
            ws.append([f"RELATÓRIO DE GESTÃO DE DOCUMENTOS — DNE/MIREME — {periodo}"])
            ws["A1"].font = Font(bold=True, size=14)
            ws.append([])

            stats = self.db.get_relatorio_stats(ano=ano, mes=mes)
            ws.append(["Indicador", "Valor"])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            kpis = [
                (f"Total Recebidos ({periodo})",  stats['total_recebidos']),
                ("Total Respondidos",             stats['total_respondidos']),
                ("Taxa de Cumprimento (%)",       f"{stats['taxa_cumprimento']}%"),
                ("Total Fora do Prazo",           stats['total_fora_prazo']),
                (f"Reuniões ({periodo})",         stats['reunioes_mes']),
                (f"Total Enviados ({periodo})",   stats['total_enviados']),
            ]
            for k, v in kpis:
                ws.append([k, v])
            ws.append([])

            ws.append([f"Desempenho por Departamento — {periodo}"])
            ws[ws.cell(ws.max_row, 1).coordinate].font = Font(bold=True)
            ws.append(["Departamento", "Total", "Dentro Prazo", "Fora Prazo", "Taxa (%)"])
            depts = self.db.get_relatorio_departamentos(ano=ano, mes=mes)
            for d in depts:
                ws.append([d['departamento'], d['total'], d['dentro_prazo'],
                            d['fora_prazo'], f"{d['taxa']}%"])

            wb.save(filepath)
            busy.fechar()
            messagebox.showinfo("Sucesso", f"Relatório exportado:\n{filepath}", parent=self)
        except Exception as e:
            busy.fechar()
            messagebox.showerror("Erro", f"Falha ao exportar:\n{e}", parent=self)

    # ------------------------------------------------------------------ exportar PDF
    def exportar_pdf(self):
        if not _carregar_matplotlib():
            messagebox.showerror("Erro", "Biblioteca matplotlib não disponível para gerar PDF.", parent=self)
            return

        ano, mes = self._get_filtros()
        periodo = self._periodo_label()
        nome_ficheiro = f"relatorio_dne_{periodo.replace(' ', '_').replace('/', '-')}.pdf"

        filepath = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialfile=nome_ficheiro,
            parent=self
        )
        if not filepath:
            return

        busy = BusyDialog(self, "A gerar relatório PDF...")
        try:
            from datetime import datetime
            from matplotlib.backends.backend_pdf import PdfPages

            stats = self.db.get_relatorio_stats(ano=ano, mes=mes)
            depts = self.db.get_relatorio_departamentos(ano=ano, mes=mes)
            remetentes = self.db.get_remetentes_frequentes(ano=ano, mes=mes)
            gerado_em = datetime.now().strftime('%d/%m/%Y %H:%M')

            with PdfPages(filepath) as pdf:
                # ---- Página 1: KPIs + Departamentos ----
                fig = plt.figure(figsize=(8.27, 11.69))
                fig.suptitle("RELATÓRIO DE GESTÃO DE DOCUMENTOS\nDNE | MIREME",
                              fontsize=16, fontweight='bold', y=0.97)
                fig.text(0.5, 0.915, f"Período: {periodo}    |    Gerado em: {gerado_em}",
                         ha='center', fontsize=10, color='gray')

                # Tabela de KPIs
                ax_kpi = fig.add_axes([0.12, 0.66, 0.76, 0.20])
                ax_kpi.axis('off')
                ax_kpi.set_title("Indicadores Chave de Desempenho", fontsize=12,
                                  fontweight='bold', loc='left', pad=10)
                kpi_rows = [
                    ["Total Recebidos", str(stats['total_recebidos'])],
                    ["Total Respondidos", str(stats['total_respondidos'])],
                    ["Taxa de Cumprimento", f"{stats['taxa_cumprimento']}%"],
                    ["Fora do Prazo", str(stats['total_fora_prazo'])],
                    ["Reuniões", str(stats['reunioes_mes'])],
                    ["Total Enviados", str(stats['total_enviados'])],
                ]
                self._tabela_pdf(ax_kpi, ["Indicador", "Valor"], kpi_rows, col_widths=[0.65, 0.35])

                # Tabela de departamentos
                ax_dep = fig.add_axes([0.08, 0.30, 0.84, 0.30])
                ax_dep.axis('off')
                ax_dep.set_title(f"Desempenho por Departamento — {periodo}", fontsize=12,
                                  fontweight='bold', loc='left', pad=10)
                dep_rows = [[d['departamento'], str(d['total']), str(d['dentro_prazo']),
                             str(d['fora_prazo']), f"{d['taxa']}%"] for d in depts]
                if not dep_rows:
                    dep_rows = [["— Sem dados para o período —", "", "", "", ""]]
                self._tabela_pdf(ax_dep,
                                 ["Departamento", "Total", "Dentro Prazo", "Fora Prazo", "Taxa (%)"],
                                 dep_rows, col_widths=[0.42, 0.145, 0.145, 0.145, 0.145])

                fig.text(0.5, 0.02, "Sistema de Gestão de Documentos — DNE/MIREME © Iazalde Jose Jeremias",
                         ha='center', fontsize=8, color='gray')
                pdf.savefig(fig)
                plt.close(fig)

                # ---- Página 2: Gráfico + Remetentes ----
                fig2 = plt.figure(figsize=(8.27, 11.69))
                fig2.suptitle("RELATÓRIO DE GESTÃO DE DOCUMENTOS\nDNE | MIREME",
                               fontsize=16, fontweight='bold', y=0.97)
                fig2.text(0.5, 0.915, f"Período: {periodo}    |    Gerado em: {gerado_em}",
                          ha='center', fontsize=10, color='gray')

                if dep_rows and depts:
                    ax_chart = fig2.add_axes([0.10, 0.60, 0.80, 0.28])
                    labels = [d['departamento'].replace("Dep. ", "").replace("Rep. ", "") for d in depts]
                    dentro = [d['dentro_prazo'] for d in depts]
                    fora = [d['fora_prazo'] for d in depts]
                    x = range(len(labels))
                    width = 0.35
                    ax_chart.bar([i - width / 2 for i in x], dentro, width, label='Dentro do Prazo', color='#27ae60')
                    ax_chart.bar([i + width / 2 for i in x], fora, width, label='Fora do Prazo', color='#c0392b')
                    ax_chart.set_xticks(list(x))
                    ax_chart.set_xticklabels(labels, rotation=20, ha='right', fontsize=8)
                    ax_chart.set_ylabel('Documentos')
                    ax_chart.set_title('Cumprimento por Departamento', fontsize=11)
                    ax_chart.legend(fontsize=8)

                ax_rem = fig2.add_axes([0.12, 0.30, 0.76, 0.22])
                ax_rem.axis('off')
                ax_rem.set_title(f"Principais Remetentes — {periodo}", fontsize=12,
                                  fontweight='bold', loc='left', pad=10)
                rem_rows = [[r['proveniencia'], str(r['total'])] for r in remetentes]
                if not rem_rows:
                    rem_rows = [["— Sem dados para o período —", ""]]
                self._tabela_pdf(ax_rem, ["Proveniência / Instituição", "Total Documentos"],
                                 rem_rows, col_widths=[0.75, 0.25])

                fig2.text(0.5, 0.02, "Sistema de Gestão de Documentos — DNE/MIREME © Iazalde Jose Jeremias",
                          ha='center', fontsize=8, color='gray')
                pdf.savefig(fig2)
                plt.close(fig2)

            busy.fechar()
            messagebox.showinfo("Sucesso", f"Relatório PDF exportado:\n{filepath}", parent=self)
        except Exception as e:
            busy.fechar()
            messagebox.showerror("Erro", f"Falha ao exportar PDF:\n{e}", parent=self)

    @staticmethod
    def _tabela_pdf(ax, col_labels, rows, col_widths=None):
        """Desenha uma tabela simples num eixo matplotlib, com cabeçalho destacado."""
        tbl = ax.table(cellText=rows, colLabels=col_labels, loc='center',
                        cellLoc='left', colWidths=col_widths)
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(8)
        tbl.scale(1, 1.4)
        for (row, col), cell in tbl.get_celld().items():
            if row == 0:
                cell.set_facecolor('#1F4E79')
                cell.set_text_props(color='white', fontweight='bold')
            else:
                cell.set_facecolor('#f0f4f8' if row % 2 == 0 else 'white')

    def on_activate(self):
        # Actualiza lista de anos disponíveis ao entrar no módulo
        anos = self.db.get_anos_disponiveis()
        self.cmb_ano.configure(values=anos)
        self.refresh()
