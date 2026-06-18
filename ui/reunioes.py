import calendar
import webbrowser
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, date
import customtkinter as ctk
from ui.widgets import DateEntry, enable_sorting, BusyDialog, enable_unsaved_changes_guard
from utils import parse_hora_local, get_meeting_datetimes, iso_to_display, display_to_iso

# Horário de expediente para agendamento de reuniões: 07:30 às 18:00
HORAS_EXPEDIENTE = []
_h, _m = 7, 30
while (_h, _m) <= (18, 0):
    HORAS_EXPEDIENTE.append(f"{_h:02d}:{_m:02d}")
    _m += 15
    if _m == 60:
        _m = 0
        _h += 1

MONTHS_PT = [
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
]
WEEKDAYS_PT = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]


class ReunioesFrame(ctk.CTkFrame):
    def __init__(self, parent, db, config):
        super().__init__(parent, corner_radius=0)
        self.db = db
        self.config = config
        self.selected_date = None
        self._cal_year = date.today().year
        self._cal_month = date.today().month

        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self._build_toolbar()
        self._build_content()
        self.refresh()

    def _build_toolbar(self):
        tb = ctk.CTkFrame(self, height=56, corner_radius=0, fg_color=("gray90", "gray20"))
        tb.grid(row=0, column=0, sticky="ew")
        tb.grid_columnconfigure(2, weight=1)

        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *a: self.refresh())
        ctk.CTkLabel(tb, text="🔍").grid(row=0, column=0, padx=(10, 2), pady=10)
        ctk.CTkEntry(tb, textvariable=self.search_var, placeholder_text="Pesquisar...",
                     width=220).grid(row=0, column=1, padx=4, pady=10)

        btn_frame = ctk.CTkFrame(tb, fg_color="transparent")
        btn_frame.grid(row=0, column=3, padx=10, pady=6, sticky="e")
        ctk.CTkButton(btn_frame, text="+ Nova", width=80, command=self.open_new,
                      fg_color="#1F4E79").pack(side="left", padx=3)
        ctk.CTkButton(btn_frame, text="✏️ Editar", width=90, command=self.open_edit,
                      fg_color="#2c6fad").pack(side="left", padx=3)
        ctk.CTkButton(btn_frame, text="🗑️ Eliminar", width=100, command=self.delete_selected,
                      fg_color="#c0392b").pack(side="left", padx=3)
        ctk.CTkButton(btn_frame, text="🔄 Todos", width=80, command=self._clear_filter,
                      fg_color="gray50").pack(side="left", padx=3)

    def _build_content(self):
        paned = ctk.CTkFrame(self, corner_radius=0)
        paned.grid(row=1, column=0, sticky="nsew")
        paned.grid_rowconfigure(0, weight=1)
        paned.grid_columnconfigure(1, weight=1)

        # Calendar panel
        self.cal_frame = ctk.CTkFrame(paned, width=260, corner_radius=8)
        self.cal_frame.grid(row=0, column=0, sticky="ns", padx=10, pady=10)
        self.cal_frame.grid_propagate(False)
        self._build_calendar(self.cal_frame)

        # List panel
        list_frame = ctk.CTkFrame(paned, corner_radius=0)
        list_frame.grid(row=0, column=1, sticky="nsew", padx=(0, 10), pady=10)
        list_frame.grid_rowconfigure(0, weight=1)
        list_frame.grid_columnconfigure(0, weight=1)
        self._build_table(list_frame)

    def _build_calendar(self, parent):
        nav = ctk.CTkFrame(parent, fg_color="transparent")
        nav.pack(fill="x", padx=5, pady=(8, 2))
        ctk.CTkButton(nav, text="◀", width=28, command=self._prev_month).pack(side="left")
        self.lbl_month = ctk.CTkLabel(nav, text="", font=ctk.CTkFont(size=13, weight="bold"))
        self.lbl_month.pack(side="left", expand=True)
        ctk.CTkButton(nav, text="▶", width=28, command=self._next_month).pack(side="right")

        header_frame = ctk.CTkFrame(parent, fg_color="transparent")
        header_frame.pack(fill="x", padx=5)
        for wd in WEEKDAYS_PT:
            ctk.CTkLabel(header_frame, text=wd, width=32, font=ctk.CTkFont(size=10, weight="bold"),
                         text_color="gray").pack(side="left", padx=1)

        self.cal_grid = ctk.CTkFrame(parent, fg_color="transparent")
        self.cal_grid.pack(fill="both", expand=True, padx=5, pady=5)
        self._render_calendar()

    def _render_calendar(self):
        for w in self.cal_grid.winfo_children():
            w.destroy()
        self.lbl_month.configure(text=f"{MONTHS_PT[self._cal_month]} {self._cal_year}")

        # Get days with meetings
        all_reunioes = self.db.get_all_reunioes()
        meeting_days = set()
        for r in all_reunioes:
            dr = r.get('data_reuniao', '')
            if dr and dr.startswith(f"{self._cal_year}-{self._cal_month:02d}"):
                try:
                    meeting_days.add(int(dr.split('-')[2]))
                except Exception:
                    pass

        today = date.today()
        cal = calendar.monthcalendar(self._cal_year, self._cal_month)
        for week in cal:
            row_frame = ctk.CTkFrame(self.cal_grid, fg_color="transparent")
            row_frame.pack(fill="x", pady=1)
            for day in week:
                if day == 0:
                    ctk.CTkLabel(row_frame, text="", width=32).pack(side="left", padx=1)
                else:
                    d = date(self._cal_year, self._cal_month, day)
                    if d == today:
                        fg = "#27ae60"
                    elif day in meeting_days:
                        fg = "#2c6fad"
                    elif d < today:
                        fg = "gray40"
                    else:
                        fg = ("gray70", "gray50")
                    btn = ctk.CTkButton(row_frame, text=str(day), width=32, height=28,
                                        fg_color=fg, hover_color="#1a4d7d",
                                        font=ctk.CTkFont(size=10),
                                        command=lambda dd=day: self._filter_by_day(dd))
                    btn.pack(side="left", padx=1)

    def _prev_month(self):
        self._cal_month -= 1
        if self._cal_month < 1:
            self._cal_month = 12
            self._cal_year -= 1
        self._render_calendar()

    def _next_month(self):
        self._cal_month += 1
        if self._cal_month > 12:
            self._cal_month = 1
            self._cal_year += 1
        self._render_calendar()

    def _filter_by_day(self, day):
        iso = f"{self._cal_year}-{self._cal_month:02d}-{day:02d}"
        self.selected_date = iso
        self.refresh()

    def _clear_filter(self):
        self.selected_date = None
        self.search_var.set("")
        self.refresh()

    def _build_table(self, parent):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Reu.Treeview", rowheight=26, font=('Segoe UI', 10))
        style.configure("Reu.Treeview.Heading", font=('Segoe UI', 10, 'bold'),
                        background="#1F4E79", foreground="white")
        style.map("Reu.Treeview", background=[("selected", "#2c6fad")])

        cols = ("id", "num_doc", "organizador", "data_conv", "assunto",
                "data_reuniao", "dias_falta", "hora", "local", "status")
        self.tree = ttk.Treeview(parent, columns=cols, show="headings",
                                 style="Reu.Treeview", selectmode="browse")

        col_config = [
            ("id", "ID", 40), ("num_doc", "Nº Doc", 160), ("organizador", "Organizador", 140),
            ("data_conv", "Data Conv.", 100), ("assunto", "Assunto", 260),
            ("data_reuniao", "Data Reunião", 100), ("dias_falta", "Dias em Falta", 90),
            ("hora", "Hora", 110), ("local", "Local", 140), ("status", "Status", 90),
        ]
        for col, heading, width in col_config:
            self.tree.heading(col, text=heading)
            self.tree.column(col, width=width, minwidth=40)

        self.tree.tag_configure("concluida", background="#d4edda")
        self.tree.tag_configure("urgente", background="#f8d7da")
        self.tree.tag_configure("em_breve", background="#fff3cd")
        self.tree.tag_configure("cancelada", background="#b0b0b0", foreground="#444444")

        vsb = ttk.Scrollbar(parent, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree.bind("<Double-1>", lambda e: self.open_edit())
        self.tree.bind("<Return>",   lambda e: self.open_edit())
        self.tree.bind("<Delete>",   lambda e: self.delete_selected())
        enable_sorting(self.tree, [c for c in cols if c != "id"])

    def refresh(self, *args):
        filters = {}
        s = self.search_var.get().strip()
        if s:
            filters['search'] = s
        if self.selected_date:
            filters['data_reuniao'] = self.selected_date

        rows = self.db.get_all_reunioes(filters)
        for item in self.tree.get_children():
            self.tree.delete(item)
        today = date.today()
        agora = datetime.now()
        for r in rows:
            dr_iso = r.get('data_reuniao', '')
            hora_local = r.get('hora_local', '')
            dias_falta = ""
            tag = ""
            status = ""

            if r.get('cancelada', 0):
                tag = "cancelada"
                status = "Cancelada"
                if dr_iso:
                    try:
                        dr = datetime.strptime(dr_iso, "%Y-%m-%d").date()
                        dias_falta = str((dr - today).days)
                    except Exception:
                        pass
            elif dr_iso:
                try:
                    dr = datetime.strptime(dr_iso, "%Y-%m-%d").date()
                    diff = (dr - today).days
                    dias_falta = str(diff)
                except Exception:
                    pass

                inicio, fim = get_meeting_datetimes(dr_iso, hora_local)
                if fim is not None and agora > fim:
                    tag = "concluida"
                    status = "Realizada"
                elif inicio is not None:
                    diff_horas = (inicio - agora).total_seconds() / 3600
                    if diff_horas <= 24:
                        tag = "urgente"
                        status = "Urgente"
                    elif diff_horas <= 48:
                        tag = "em_breve"
                        status = "Em Breve"
                    else:
                        status = "Agendada"

            hora_inicio, hora_fim, local = parse_hora_local(hora_local)
            if hora_inicio and hora_fim:
                hora_txt = f"{hora_inicio} - {hora_fim}"
            else:
                hora_txt = hora_inicio or hora_fim

            self.tree.insert("", "end", iid=str(r['id']), tags=(tag,), values=(
                r['id'], r.get('num_doc', ''), r.get('organizador', ''),
                iso_to_display(r.get('data_convocatoria', '')),
                r.get('assunto', ''),
                iso_to_display(dr_iso),
                dias_falta,
                hora_txt,
                local,
                status,
            ))
        self._render_calendar()

    def _get_selected_id(self):
        sel = self.tree.selection()
        if not sel:
            return None
        return int(sel[0])

    def open_new(self):
        ReuniaoForm(self, self.db, self.config, None, self.refresh)

    def open_edit(self):
        rid = self._get_selected_id()
        if rid is None:
            messagebox.showwarning("Aviso", "Seleccione uma reunião.", parent=self)
            return
        ReuniaoForm(self, self.db, self.config, rid, self.refresh)

    def delete_selected(self):
        rid = self._get_selected_id()
        if rid is None:
            messagebox.showwarning("Aviso", "Seleccione uma reunião.", parent=self)
            return
        reuniao = self.db.get_reuniao(rid)
        assunto = reuniao.get('assunto', str(rid)) if reuniao else str(rid)
        if messagebox.askyesno("Confirmar", f"Eliminar a reunião:\n{assunto}?", parent=self):
            self.db.delete_reuniao(rid)
            self.refresh()

    def exportar(self):
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="reunioes.xlsx",
            parent=self
        )
        if not filepath:
            return
        busy = BusyDialog(self, "A exportar para Excel...")
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            rows = self.db.get_all_reunioes()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reuniões"
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="1F4E79")
            headers = ['ID', 'Nº Doc', 'Organizador', 'Data Conv.', 'Assunto',
                       'Data Reunião', 'Hora/Local', 'Participantes', 'Decisões']
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            for r in rows:
                ws.append([r.get('id'), r.get('num_doc'), r.get('organizador'),
                           r.get('data_convocatoria'), r.get('assunto'),
                           r.get('data_reuniao'), r.get('hora_local'),
                           r.get('participantes'), r.get('decisoes')])
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 20
            wb.save(filepath)
            busy.fechar()
            messagebox.showinfo("Sucesso", f"Exportado para:\n{filepath}", parent=self)
        except Exception as e:
            busy.fechar()
            messagebox.showerror("Erro", f"Falha ao exportar:\n{e}", parent=self)

    def focus_search(self):
        try:
            for w in self.winfo_children():
                if hasattr(w, 'winfo_children'):
                    for child in w.winfo_children():
                        if isinstance(child, ctk.CTkEntry):
                            child.focus_set()
                            return
        except Exception:
            pass

    def on_activate(self):
        self.refresh()


class ReuniaoForm(ctk.CTkToplevel):
    def __init__(self, parent, db, config, record_id, callback):
        super().__init__(parent)
        self.db = db
        self.config = config
        self.record_id = record_id
        self.callback = callback
        self.title("Nova Reunião" if not record_id else "Editar Reunião")
        self.geometry("720x660")
        self.grab_set()
        self._vars = {}
        self._nomes = [r['nome'] for r in db.get_all_contactos()]
        self._build_form()
        if record_id:
            self._load_data(record_id)

        self.bind("<Control-s>", lambda e: self._save())
        enable_unsaved_changes_guard(self)

    def _lbl_entry(self, parent, row, col, label, var_key, width=280):
        ctk.CTkLabel(parent, text=label, anchor="e", width=130).grid(
            row=row, column=col * 2, padx=(10, 4), pady=6, sticky="e")
        var = tk.StringVar()
        entry = ctk.CTkEntry(parent, textvariable=var, width=width)
        entry.grid(row=row, column=col * 2 + 1, padx=(0, 10), pady=6, sticky="w")
        self._vars[var_key] = var
        return entry

    def _build_form(self):
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        scroll = ctk.CTkScrollableFrame(self)
        scroll.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        f = scroll

        self._lbl_entry(f, 0, 0, "Nº Documento", "num_doc", 280)

        ctk.CTkLabel(f, text="Organizador", anchor="e", width=130).grid(row=0, column=2, padx=(10, 4), pady=6, sticky="e")
        self._vars['organizador'] = tk.StringVar()
        ctk.CTkComboBox(f, values=self._nomes, variable=self._vars['organizador'],
                        width=240).grid(row=0, column=3, padx=(0, 10), pady=6, sticky="w")

        ctk.CTkLabel(f, text="Assunto *", anchor="e", width=130).grid(row=1, column=0, padx=(10, 4), pady=6, sticky="e")
        self._vars['assunto'] = tk.StringVar()
        ctk.CTkEntry(f, textvariable=self._vars['assunto'], width=480).grid(
            row=1, column=1, columnspan=3, padx=(0, 10), pady=6, sticky="w")

        ctk.CTkLabel(f, text="Data Convocatória", anchor="e", width=130).grid(
            row=2, column=0, padx=(10, 4), pady=6, sticky="e")
        self._vars['data_convocatoria'] = tk.StringVar()
        DateEntry(f, textvariable=self._vars['data_convocatoria'], width=120).grid(
            row=2, column=1, padx=(0, 10), pady=6, sticky="w")

        ctk.CTkLabel(f, text="Data Reunião", anchor="e", width=130).grid(
            row=2, column=2, padx=(10, 4), pady=6, sticky="e")
        self._vars['data_reuniao'] = tk.StringVar()
        dr_frame = ctk.CTkFrame(f, fg_color="transparent")
        dr_frame.grid(row=2, column=3, padx=(0, 10), pady=6, sticky="w")
        DateEntry(dr_frame, textvariable=self._vars['data_reuniao'], width=120).pack(side="left")
        self._vars['cancelada'] = tk.IntVar()
        ctk.CTkCheckBox(dr_frame, text="Cancelada", variable=self._vars['cancelada'],
                        text_color="#c0392b", fg_color="#c0392b",
                        hover_color="#e74c3c").pack(side="left", padx=(14, 0))

        ctk.CTkLabel(f, text="Hora (Início - Fim)", anchor="e", width=130).grid(
            row=3, column=0, padx=(10, 4), pady=6, sticky="e")
        hora_frame = ctk.CTkFrame(f, fg_color="transparent")
        hora_frame.grid(row=3, column=1, padx=(0, 10), pady=6, sticky="w")
        self._vars['hora_inicio'] = tk.StringVar()
        self._vars['hora_fim'] = tk.StringVar()
        ctk.CTkComboBox(hora_frame, values=HORAS_EXPEDIENTE, variable=self._vars['hora_inicio'],
                        width=100).pack(side="left")
        ctk.CTkLabel(hora_frame, text=" às ").pack(side="left", padx=4)
        ctk.CTkComboBox(hora_frame, values=HORAS_EXPEDIENTE, variable=self._vars['hora_fim'],
                        width=100).pack(side="left")

        ctk.CTkLabel(f, text="Local", anchor="e", width=130).grid(
            row=3, column=2, padx=(10, 4), pady=6, sticky="e")
        self._vars['local'] = tk.StringVar()
        ctk.CTkEntry(f, textvariable=self._vars['local'], width=240).grid(
            row=3, column=3, padx=(0, 10), pady=6, sticky="w")

        ctk.CTkLabel(f, text="Link Convocatória", anchor="e", width=130).grid(row=4, column=0, padx=(10, 4), pady=6, sticky="e")
        link_frame = ctk.CTkFrame(f, fg_color="transparent")
        link_frame.grid(row=4, column=1, columnspan=3, padx=(0, 10), pady=6, sticky="w")
        self._vars['link_convocatoria'] = tk.StringVar()
        ctk.CTkEntry(link_frame, textvariable=self._vars['link_convocatoria'], width=360).pack(side="left", padx=(0, 6))
        ctk.CTkButton(link_frame, text="🔗 Abrir", width=80,
                      command=self._open_link).pack(side="left")

        ctk.CTkLabel(f, text="Participantes", anchor="e", width=130).grid(row=5, column=0, padx=(10, 4), pady=6, sticky="ne")
        self._participantes_text = ctk.CTkTextbox(f, width=480, height=70)
        self._participantes_text.grid(row=5, column=1, columnspan=3, padx=(0, 10), pady=6, sticky="w")

        ctk.CTkLabel(f, text="Contactos", anchor="e", width=130).grid(row=6, column=0, padx=(10, 4), pady=6, sticky="ne")
        self._contactos_text = ctk.CTkTextbox(f, width=480, height=50)
        self._contactos_text.grid(row=6, column=1, columnspan=3, padx=(0, 10), pady=6, sticky="w")

        ctk.CTkLabel(f, text="Decisões", anchor="e", width=130).grid(row=7, column=0, padx=(10, 4), pady=6, sticky="ne")
        self._decisoes_text = ctk.CTkTextbox(f, width=480, height=70)
        self._decisoes_text.grid(row=7, column=1, columnspan=3, padx=(0, 10), pady=6, sticky="w")

        ctk.CTkLabel(f, text="Ficheiro", anchor="e", width=130).grid(row=8, column=0, padx=(10, 4), pady=6, sticky="e")
        ff = ctk.CTkFrame(f, fg_color="transparent")
        ff.grid(row=8, column=1, columnspan=3, padx=(0, 10), pady=6, sticky="w")
        self._vars['ficheiro_path'] = tk.StringVar()
        ctk.CTkEntry(ff, textvariable=self._vars['ficheiro_path'], width=360).pack(side="left", padx=(0, 6))
        ctk.CTkButton(ff, text="📂 Procurar", width=100, command=self._pick_file).pack(side="left")

        btn_frame = ctk.CTkFrame(self)
        btn_frame.grid(row=1, column=0, pady=10)
        ctk.CTkButton(btn_frame, text="💾 Guardar", width=120, command=self._save,
                      fg_color="#1F4E79").pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="📋 Copiar Tudo", width=130, command=self._copiar_tudo,
                      fg_color="#5a6e8a").pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="❌ Cancelar", width=100, command=self.destroy,
                      fg_color="gray50").pack(side="left", padx=10)

    def _open_link(self):
        link = self._vars['link_convocatoria'].get().strip()
        if link:
            webbrowser.open(link)

    def _pick_file(self):
        path = filedialog.askopenfilename(parent=self)
        if path:
            self._vars['ficheiro_path'].set(path)

    def _load_data(self, rid):
        r = self.db.get_reuniao(rid)
        if not r:
            return
        for key in ('num_doc', 'organizador', 'assunto',
                    'link_convocatoria', 'ficheiro_path'):
            if key in self._vars and r.get(key):
                self._vars[key].set(r[key])
        self._vars['data_convocatoria'].set(iso_to_display(r.get('data_convocatoria', '')))
        self._vars['data_reuniao'].set(iso_to_display(r.get('data_reuniao', '')))
        self._vars['cancelada'].set(int(r.get('cancelada', 0) or 0))

        # Compatibilidade: campo antigo "hora_local" tipo "12:20-17:15 Local X"
        hora_inicio, hora_fim, local = parse_hora_local(r.get('hora_local', ''))
        if hora_inicio:
            self._vars['hora_inicio'].set(hora_inicio)
        if hora_fim:
            self._vars['hora_fim'].set(hora_fim)
        if local:
            self._vars['local'].set(local)
        for widget, key in [(self._participantes_text, 'participantes'),
                            (self._contactos_text, 'contactos'),
                            (self._decisoes_text, 'decisoes')]:
            val = r.get(key, '') or ''
            widget.delete("1.0", "end")
            widget.insert("1.0", val)

    def _montar_hora_local(self):
        ini = self._vars['hora_inicio'].get().strip()
        fim = self._vars['hora_fim'].get().strip()
        local = self._vars['local'].get().strip()
        if ini and fim:
            hora = f"{ini} - {fim}"
        else:
            hora = ini or fim
        partes = [p for p in (hora, local) if p]
        return "  ".join(partes)

    def _save(self):
        assunto = self._vars['assunto'].get().strip()
        if not assunto:
            messagebox.showerror("Erro", "Assunto é obrigatório.", parent=self)
            return

        ini = self._vars['hora_inicio'].get().strip()
        fim = self._vars['hora_fim'].get().strip()
        if ini and (ini < "07:30" or ini > "18:00"):
            messagebox.showerror("Erro", "A hora de início deve estar entre 07:30 e 18:00.", parent=self)
            return
        if fim and (fim < "07:30" or fim > "18:00"):
            messagebox.showerror("Erro", "A hora de término deve estar entre 07:30 e 18:00.", parent=self)
            return

        data = {
            'num_doc': self._vars['num_doc'].get().strip(),
            'organizador': self._vars['organizador'].get().strip(),
            'assunto': assunto,
            'data_convocatoria': display_to_iso(self._vars['data_convocatoria'].get().strip()),
            'data_reuniao': display_to_iso(self._vars['data_reuniao'].get().strip()),
            'hora_local': self._montar_hora_local(),
            'link_convocatoria': self._vars['link_convocatoria'].get().strip(),
            'participantes': self._participantes_text.get("1.0", "end").strip(),
            'contactos': self._contactos_text.get("1.0", "end").strip(),
            'decisoes': self._decisoes_text.get("1.0", "end").strip(),
            'ficheiro_path': self._vars['ficheiro_path'].get().strip(),
            'cancelada': int(self._vars['cancelada'].get()),
        }
        try:
            if self.record_id:
                self.db.update_reuniao(self.record_id, data)
            else:
                self.db.insert_reuniao(data)
            self.callback()
            self.destroy()
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao guardar:\n{e}", parent=self)

    def _copiar_tudo(self):
        linhas = [
            f"Nº Documento: {self._vars['num_doc'].get().strip()}",
            f"Organizador: {self._vars['organizador'].get().strip()}",
            f"Assunto: {self._vars['assunto'].get().strip()}",
            f"Data Convocatória: {self._vars['data_convocatoria'].get().strip()}",
            f"Data Reunião: {self._vars['data_reuniao'].get().strip()}",
            f"Estado: {'Cancelada' if self._vars['cancelada'].get() else 'Activa'}",
            f"Hora Início: {self._vars['hora_inicio'].get().strip()}",
            f"Hora Fim: {self._vars['hora_fim'].get().strip()}",
            f"Local: {self._vars['local'].get().strip()}",
            f"Link Convocatória: {self._vars['link_convocatoria'].get().strip()}",
            f"Participantes:\n{self._participantes_text.get('1.0', 'end').strip()}",
            f"Contactos:\n{self._contactos_text.get('1.0', 'end').strip()}",
            f"Decisões:\n{self._decisoes_text.get('1.0', 'end').strip()}",
            f"Ficheiro: {self._vars['ficheiro_path'].get().strip()}",
        ]
        texto = "\n".join(linhas)
        self.clipboard_clear()
        self.clipboard_append(texto)
        messagebox.showinfo("Copiado", "Conteúdo copiado para a área de transferência.", parent=self)
